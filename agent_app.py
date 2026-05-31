"""
Agent App — v1.0
----------------
1099 Pre-Reconciliation orchestration with three engines:

    * Rule-based only         — deterministic pipeline, no AI
    * AI-assisted             — single Claude agent with MCP tools, sequential
    * Multi-agent validation  — one agent per PDF (concurrent) + deterministic
                                cross-statement validation downstream

v1.0 changes from v0.6:
    * AI Validation Narrative call removed entirely. Deterministic findings
      from validation_engine.py are now the sole source of cross-statement
      output; rendering is the consumer's responsibility (web UI / workbook).
    * Per-agent prose summary suppressed via system prompt — the agent only
      executes tools, no closing narrative, no per-vendor commentary, no
      filesystem path leakage. The `agent_narrative` field is no longer
      returned.
    * `validation_model` parameter removed from all engine signatures.
    * `language` parameter retained as a no-op pass-through; reserved for
      future workbook/report localization in v1.1.

Model strategy (updated in v1.2):
    Extraction model:  Sonnet by default. Opus available as advanced
                       override via extraction_model parameter. Haiku
                       removed — not appropriate for structured PDF
                       extraction and accounting analysis.
    Validation model:  Removed in v1.0. The deterministic engine has no model.

v1.2 — Transaction Classifier integration in rule-based engine
---------------------------------------------------------------
_run_rule_based previously had a subtle dual-extraction issue: it called
run_pipeline() to generate the Excel (which classifies via pipeline.py),
then called extract_transactions() a second time to build the JSON response
payload — and that second pass did NOT classify. Result: the per-statement
Excel was correct but the UI Workspace KPIs were not.

The fix wires the classifier into the second-pass code so both paths are
consistent. The JSON response now reflects classified data: only
include_for_1099 == True rows contribute to vendor totals, transaction
counts, and over-$600 candidate counts.

Status taxonomy:
    success            — all pipeline steps completed cleanly
    partial            — some steps completed, vendors extracted
    failed_rate_limit  — 429 from Anthropic API
    failed_extraction  — PDF parsing returned 0 transactions
    failed_other       — any other exception

Failure handling:
    asyncio.gather(return_exceptions=True) ensures one agent failure
    never breaks the others.

Filename handling (v1.0):
    The agent layer uses uuid-based paths (whatever server.py hands it).
    `statement_label` in returned dicts is `Path(pdf_path).name` — i.e., the
    saved-uuid filename. server.py resolves to original_filename at the
    response boundary using its filename_map. The agent layer does not need
    to know original filenames.
"""

import os
import asyncio
import sys
from pathlib import Path

# Ensure backend modules are importable
sys.path.insert(0, str(Path(__file__).parent / "backend"))

from claude_agent_sdk import (
    ClaudeAgentOptions,
    ClaudeSDKClient,
    AssistantMessage,
    TextBlock,
    ResultMessage,
)

from agent_tools import (
    reconciliation_server,
    reset_session,
    get_session_summaries,
    get_session_transactions,
    get_session_normalized,
)


# ---------------------------------------------------------------------------
# Model constants — extraction default updated in v1.2
# ---------------------------------------------------------------------------
# v1.2: Default extraction model changed from Haiku to Sonnet. Mentor
# feedback: Haiku is appropriate for lightweight conversational tasks but
# is not the right model class for structured PDF extraction and accounting
# analysis. Sonnet is the recommended default. Opus remains available as an
# advanced override (selectable in the frontend Advanced panel).

EXTRACTION_MODEL_DEFAULT = "claude-sonnet-4-6"


# ---------------------------------------------------------------------------
# System prompt — silenced in v1.0
# ---------------------------------------------------------------------------
# The agent's job is to call tools, not to narrate. No closing summary, no
# per-vendor commentary, no filesystem paths. Tool outputs feed downstream;
# the deterministic engines and the workbook do the user-facing rendering.

SYSTEM_PROMPT = """You are a tax preparation assistant. Your only job is to \
execute these tools in order on the provided PDF:

1. extract_pdf_transactions
2. load_vendor_list (only if a CSV path was provided)
3. normalize_vendors
4. aggregate_by_vendor
5. generate_excel_report

Rules:
- Call each tool exactly once, in order.
- Do not produce any narrative output, summary, recommendation, or commentary.
- Do not list flagged vendors, file paths, or interpretive notes.
- The Excel file is the deliverable; the master workbook downstream will
  consolidate findings. Your text output is not rendered to the user.
- After generate_excel_report returns, stop. No closing message.
"""


# ---------------------------------------------------------------------------
# Status helpers
# ---------------------------------------------------------------------------

def _classify_error(exc: Exception) -> tuple[str, str]:
    """Classify an exception into (status, message)."""
    msg = str(exc)
    msg_lower = msg.lower()
    if "rate_limit" in msg_lower or "429" in msg or "rate limit" in msg_lower:
        return "failed_rate_limit", (
            "API rate limit exceeded. Try a smaller/faster model or fewer "
            "concurrent agents. " + msg[:200]
        )
    if "credit" in msg_lower and "balance" in msg_lower:
        return "failed_other", "API credit balance too low: " + msg[:200]
    return "failed_other", msg[:300]


def _serialize_summaries(summaries: list) -> list[dict]:
    """Convert VendorSummary objects to plain dicts for JSON serialization."""
    out = []
    for s in summaries:
        out.append({
            "canonical_name":      s.canonical_name,
            "entity_type":         s.entity_type,
            "total_amount":        round(s.total_amount, 2),
            "transaction_count":   s.transaction_count,
            "first_payment_date":  s.first_payment_date,
            "last_payment_date":   s.last_payment_date,
            "match_confidence":    s.match_confidence,
            "needs_review":        s.needs_review,
            "review_reasons":      list(s.review_reasons),
            "raw_name_variants":   list(s.raw_name_variants),
        })
    return out


def _serialize_transactions(transactions: list, normalized: list) -> list[dict]:
    """Convert Transaction objects to plain dicts.

    v1.2: surfaces transaction_type and include_for_1099 if the classifier
    has run. Old/unclassified Transaction objects fall back to defaults.
    """
    out = []
    for t, n in zip(transactions, normalized or [None] * len(transactions)):
        out.append({
            "date":              getattr(t, "date", ""),
            "raw_description":   getattr(t, "description", ""),
            "amount":            round(getattr(t, "amount", 0.0), 2),
            "canonical_name":    n.canonical_name if n else "",
            "excluded":          bool(n.excluded) if n else False,
            "exclusion_reason":  getattr(n, "exclusion_reason", "") if n else
                                 getattr(t, "exclusion_reason", ""),
            # v1.2: classifier outputs (default vendor_payment / True for
            # backward-compat with any code path that bypasses the classifier)
            "transaction_type":  getattr(t, "transaction_type", "vendor_payment"),
            "include_for_1099":  getattr(t, "include_for_1099", True),
            "review_required":   getattr(t, "review_required", False),
        })
    return out


# ---------------------------------------------------------------------------
# Single-agent runner
# ---------------------------------------------------------------------------

async def run_single_agent(
    pdf_path: str,
    output_path: str,
    vendor_list_path: str | None = None,
    extraction_model: str = EXTRACTION_MODEL_DEFAULT,
    language: str = "English",
    session_id: str | None = None,
) -> dict:
    """
    Run one agent on one PDF. Always returns a dict with `status`, never raises.

    Note: `language` is accepted for forward compatibility (v1.1 workbook
    localization) but does not affect output in v1.0. The agent never
    produces narrative text in v1.0, so language has nothing to translate.
    """
    label = Path(pdf_path).name   # uuid-based filename; resolution happens in server.py
    reset_session(session_id)

    task = f"""PDF statement: {pdf_path}
Output Excel: {output_path}"""
    if vendor_list_path:
        task += f"\nKnown vendor list CSV: {vendor_list_path}"
    task += "\n\nExecute the tools in order. Do not produce any text output."

    options = ClaudeAgentOptions(
        system_prompt=SYSTEM_PROMPT,
        model=extraction_model,
        mcp_servers={"reconciliation": reconciliation_server},
        allowed_tools=[
            "mcp__reconciliation__extract_pdf_transactions",
            "mcp__reconciliation__load_vendor_list",
            "mcp__reconciliation__normalize_vendors",
            "mcp__reconciliation__aggregate_by_vendor",
            "mcp__reconciliation__generate_excel_report",
            "mcp__reconciliation__get_review_items",
        ],
        permission_mode="acceptEdits",
    )

    print("=" * 70)
    print(f"🤖 Agent: {label}  (model: {extraction_model})")
    print("=" * 70)

    total_cost = 0.0
    tool_calls = 0
    extraction_confidence = 0.85
    status = "success"
    error_message = None

    try:
        async with ClaudeSDKClient(options=options) as client:
            await client.query(task)
            async for message in client.receive_response():
                if isinstance(message, AssistantMessage):
                    for block in message.content:
                        if isinstance(block, TextBlock):
                            # Print to stdout for operational visibility, but
                            # do NOT capture or return — narrative is suppressed
                            # in v1.0. Any text the model produces despite the
                            # system prompt is logged for debugging only.
                            if block.text.strip():
                                print(f"[{label}] (suppressed text): {block.text[:120]}")
                        elif hasattr(block, "name"):
                            tool_calls += 1
                            tname = block.name.replace("mcp__reconciliation__", "")
                            print(f"[{label}]   🔧 Tool #{tool_calls}: {tname}")
                if isinstance(message, ResultMessage):
                    if hasattr(message, "total_cost_usd") and message.total_cost_usd:
                        total_cost = message.total_cost_usd
    except Exception as e:
        status, error_message = _classify_error(e)
        print(f"[{label}] ✗ {status}: {error_message}")

    # Pull session state
    summaries    = get_session_summaries(session_id)
    transactions = get_session_transactions(session_id)
    normalized   = get_session_normalized(session_id)

    # ── Excel-fallback for vendor summaries (unchanged from v0.6) ──
    if not summaries and Path(output_path).exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_path, data_only=True)
            ws = wb["Vendor Summary"]

            from dataclasses import dataclass, field
            from typing import Optional

            @dataclass
            class _VendorFromExcel:
                canonical_name: str
                entity_type: Optional[str]
                total_amount: float
                transaction_count: int
                first_payment_date: Optional[str] = None
                last_payment_date: Optional[str] = None
                match_confidence: float = 1.0
                needs_review: bool = False
                review_reasons: list = field(default_factory=list)
                raw_name_variants: list = field(default_factory=list)

            recovered = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if row[0] is None or str(row[0]).strip() == "" or str(row[0]) == "TOTAL":
                    break
                recovered.append(_VendorFromExcel(
                    canonical_name=str(row[0]),
                    entity_type=str(row[1]) if row[1] else None,
                    total_amount=float(row[2] or 0),
                    transaction_count=int(row[3] or 0),
                    first_payment_date=str(row[4]) if row[4] else None,
                    last_payment_date=str(row[5]) if row[5] else None,
                    match_confidence=float(row[6] or 1.0) if isinstance(row[6], (int, float)) else 1.0,
                    needs_review=row[10] == "YES" if len(row) > 10 else False,
                ))

            if recovered:
                summaries = recovered
                print(f"[{label}] ℹ Recovered {len(recovered)} vendors from Excel file")
        except Exception as e:
            print(f"[{label}] ⚠ Could not read Excel fallback: {e}")

    # ── Excel-fallback for transactions (unchanged from v0.6) ──
    if not transactions and Path(output_path).exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_path, data_only=True)
            if "Transactions" in wb.sheetnames:
                tx_ws = wb["Transactions"]
                from dataclasses import dataclass
                from typing import Optional

                @dataclass
                class _TxnFromExcel:
                    date: str
                    description: str
                    amount: float
                    source: str = "bank"

                @dataclass
                class _NormFromExcel:
                    canonical_name: str
                    entity_type: Optional[str]
                    excluded: bool = False
                    exclusion_reason: str = ""

                recovered_txns = []
                recovered_norms = []
                for row in tx_ws.iter_rows(min_row=4, values_only=True):
                    if row[0] is None or str(row[0]).strip() == "":
                        continue
                    recovered_txns.append(_TxnFromExcel(
                        date=str(row[0] or ""),
                        description=str(row[1] or ""),
                        amount=float(row[4] or 0),
                        source=str(row[5] or "bank"),
                    ))
                    recovered_norms.append(_NormFromExcel(
                        canonical_name=str(row[2] or ""),
                        entity_type=str(row[3]) if row[3] else None,
                    ))

                if recovered_txns:
                    transactions = recovered_txns
                    normalized   = recovered_norms
                    print(f"[{label}] ℹ Recovered {len(recovered_txns)} transactions from Excel file")
        except Exception as e:
            print(f"[{label}] ⚠ Could not read transactions fallback: {e}")

    # Determine final status
    if status == "success":
        if not summaries:
            status = "failed_extraction"
            error_message = "Pipeline completed but no vendors were extracted"
        elif tool_calls < 4:
            status = "partial"
            error_message = f"Only {tool_calls} of expected 4-5 tool calls completed"

    print(f"[{label}] {status.upper()}: {tool_calls} calls, ${total_cost:.4f}")

    return {
        "statement_label":       label,
        "pdf_path":              pdf_path,
        "output_path":           output_path,
        "status":                status,
        "error_message":         error_message,
        "vendors":               _serialize_summaries(summaries),
        "transactions":          _serialize_transactions(transactions, normalized),
        "tool_calls":            tool_calls,
        "cost_usd":              round(total_cost, 4),
        "extraction_confidence": extraction_confidence,
        "model":                 extraction_model,
        "session_id":            session_id,
    }


# ---------------------------------------------------------------------------
# Top-level orchestrator: runs the chosen engine
# ---------------------------------------------------------------------------

async def run_engine(
    engine: str,
    pdf_paths: list[str],
    output_dir: str,
    vendor_list_path: str | None = None,
    extraction_model: str = EXTRACTION_MODEL_DEFAULT,
    language: str = "English",
) -> dict:
    """
    Dispatch to the right engine. Returns a normalized response.

    engine: "rule_based" | "ai_assisted" | "multi_agent"

    Note: `language` is a no-op in v1.0. Carried through for v1.1 workbook
    localization.
    """
    if not pdf_paths:
        return {"success": False, "error": "No PDFs provided"}

    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)

    if engine == "rule_based":
        return await _run_rule_based(pdf_paths, output_dir_path, vendor_list_path)

    if engine == "ai_assisted":
        return await _run_ai_assisted(
            pdf_paths, output_dir_path, vendor_list_path,
            extraction_model, language,
        )

    if engine == "multi_agent":
        return await _run_multi_agent(
            pdf_paths, output_dir_path, vendor_list_path,
            extraction_model, language,
        )

    return {"success": False, "error": f"Unknown engine: {engine}"}


# ---------------------------------------------------------------------------
# Engine: rule-based (deterministic, no AI)
# ---------------------------------------------------------------------------

async def _run_rule_based(
    pdf_paths: list[str],
    output_dir: Path,
    vendor_list_path: str | None,
) -> dict:
    """Run deterministic pipeline on each PDF — sequential, no AI.

    v1.2: After the duplicate extract_transactions() call below (used to
    build the JSON response payload), the classifier is now invoked so
    that the response reflects only include_for_1099 == True rows. Without
    this, the per-statement Excel files would be correct (built by
    run_pipeline) but the UI Workspace KPIs would still show pre-classifier
    counts. Both paths must classify identically to give the UI consistent
    numbers.
    """
    from pipeline import run_pipeline

    agent_outputs = []
    for i, pdf_path in enumerate(pdf_paths):
        label = Path(pdf_path).name
        out_path = str(output_dir / f"statement_{i+1}_{Path(pdf_path).stem}.xlsx")

        try:
            stats = run_pipeline(
                pdf_path=pdf_path,
                output_path=out_path,
                vendor_list_path=vendor_list_path,
                verbose=False,
            )
            from pdf_extractor import extract_transactions
            from vendor_normalizer import normalize_vendor
            from transaction_aggregator import aggregate_by_vendor
            from transaction_classifier import (
                classify_transactions,
                filter_for_aggregation,
            )

            extr = extract_transactions(pdf_path)

            # ── v1.2: classify before normalize/aggregate, mirroring pipeline.py ──
            # Tags every Transaction with transaction_type, include_for_1099,
            # review_required, exclusion_reason. Filter to the included subset
            # so vendor totals reflect only vendor-payment rows.
            classify_transactions(extr.transactions)
            included_txns = filter_for_aggregation(extr.transactions)

            known = []
            if vendor_list_path:
                import csv
                with open(vendor_list_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    known = [r[0].strip() for r in reader if r and r[0].strip()]

            normalized = [
                normalize_vendor(t.description, known)
                for t in included_txns
            ]
            summaries = aggregate_by_vendor(included_txns, normalized)

            agent_outputs.append({
                "statement_label":       label,
                "pdf_path":              pdf_path,
                "output_path":           out_path,
                "status":                "success",
                "error_message":         None,
                "vendors":               _serialize_summaries(summaries),
                # Serialize the FULL classified list (including excluded rows)
                # so the UI/Excel can show what was filtered out and why.
                # Aggregation totals already reflect only included rows via
                # `summaries`.
                "transactions":          _serialize_transactions(
                                              extr.transactions,
                                              # Pad normalized so zip aligns
                                              # with the full transaction list:
                                              # included rows have NormalizedVendor,
                                              # excluded rows get None.
                                              _pad_normalized_to_full_list(
                                                  extr.transactions,
                                                  included_txns,
                                                  normalized,
                                              ),
                                          ),
                "tool_calls":            0,
                "cost_usd":              0.0,
                "extraction_confidence": stats.get("confidence", 0.85),
                "model":                 "rule-based",
                "session_id":            None,
            })
        except Exception as e:
            status, msg = _classify_error(e)
            agent_outputs.append({
                "statement_label":       label,
                "pdf_path":              pdf_path,
                "output_path":           out_path,
                "status":                status,
                "error_message":         msg,
                "vendors":               [],
                "transactions":          [],
                "tool_calls":            0,
                "cost_usd":              0.0,
                "extraction_confidence": 0.0,
                "model":                 "rule-based",
                "session_id":            None,
            })

    return {
        "success":        True,
        "engine":         "rule_based",
        "agent_outputs":  agent_outputs,
        "total_cost_usd": 0.0,
    }


def _pad_normalized_to_full_list(
    full_txns: list,
    included_txns: list,
    normalized_for_included: list,
) -> list:
    """
    v1.2 helper — produce a `normalized` list aligned with the FULL transaction
    list (including excluded rows). Excluded rows get None as their normalized
    entry, so _serialize_transactions's zip() walks both lists together
    without losing the excluded rows from the response payload.

    This keeps the response shape stable: the frontend sees one row per
    extracted transaction, with `excluded` / `transaction_type` /
    `include_for_1099` flags surfacing why each row was treated the way
    it was.
    """
    # Build a fast lookup from id(included_txn) → normalized_entry
    norm_by_id = {
        id(t): n for t, n in zip(included_txns, normalized_for_included)
    }
    return [norm_by_id.get(id(t)) for t in full_txns]


# ---------------------------------------------------------------------------
# Engine: AI-assisted (single agent processes each PDF sequentially)
# ---------------------------------------------------------------------------

async def _run_ai_assisted(
    pdf_paths: list[str],
    output_dir: Path,
    vendor_list_path: str | None,
    extraction_model: str,
    language: str,
) -> dict:
    """One AI agent processes each PDF sequentially."""
    agent_outputs = []
    for i, pdf_path in enumerate(pdf_paths):
        out_path = str(output_dir / f"statement_{i+1}_{Path(pdf_path).stem}.xlsx")
        sid = f"single_{i}_{Path(pdf_path).stem}"
        result = await run_single_agent(
            pdf_path=pdf_path,
            output_path=out_path,
            vendor_list_path=vendor_list_path,
            extraction_model=extraction_model,
            language=language,
            session_id=sid,
        )
        agent_outputs.append(result)

    total_cost = sum(o.get("cost_usd", 0) for o in agent_outputs)

    return {
        "success":        True,
        "engine":         "ai_assisted",
        "agent_outputs":  agent_outputs,
        "total_cost_usd": round(total_cost, 4),
    }


# ---------------------------------------------------------------------------
# Engine: Multi-agent (one agent per PDF concurrently)
# ---------------------------------------------------------------------------

async def _run_multi_agent(
    pdf_paths: list[str],
    output_dir: Path,
    vendor_list_path: str | None,
    extraction_model: str,
    language: str,
) -> dict:
    """One agent per PDF concurrently. Cross-statement validation is
    performed downstream by validation_engine.py — no narrative call."""
    tasks = []
    for i, pdf_path in enumerate(pdf_paths):
        sid = f"multi_{i}_{Path(pdf_path).stem}"
        out_path = str(output_dir / f"statement_{i+1}_{Path(pdf_path).stem}.xlsx")
        tasks.append(run_single_agent(
            pdf_path=pdf_path,
            output_path=out_path,
            vendor_list_path=vendor_list_path,
            extraction_model=extraction_model,
            language=language,
            session_id=sid,
        ))

    print(f"\n🚀 Launching {len(tasks)} agents concurrently (model: {extraction_model})\n")
    raw = await asyncio.gather(*tasks, return_exceptions=True)

    agent_outputs = []
    for r, pdf_path in zip(raw, pdf_paths):
        if isinstance(r, Exception):
            status, msg = _classify_error(r)
            agent_outputs.append({
                "statement_label":       Path(pdf_path).name,
                "pdf_path":              pdf_path,
                "output_path":           "",
                "status":                status,
                "error_message":         msg,
                "vendors":               [],
                "transactions":          [],
                "tool_calls":            0,
                "cost_usd":              0.0,
                "extraction_confidence": 0.0,
                "model":                 extraction_model,
                "session_id":            None,
            })
        else:
            agent_outputs.append(r)

    total_cost = sum(o.get("cost_usd", 0) for o in agent_outputs)

    return {
        "success":        True,
        "engine":         "multi_agent",
        "agent_outputs":  agent_outputs,
        "total_cost_usd": round(total_cost, 4),
    }
