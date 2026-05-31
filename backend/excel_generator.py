"""
Excel Output Generator
----------------------
Produces the per-statement Excel workbook with three sheets:

    1. "Vendor Summary"  — aggregated view, one row per canonical vendor
    2. "Transactions"    — full transaction detail with normalization metadata
    3. "Summary Stats"   — totals, counts, and review-needed flags

v1.3 — PDF Skill engine support:
    When the PDF Skill engine is used, transactions arrive with extra
    classifier fields populated (transaction_type, include_for_1099,
    exclusion_reason, review_required). This generator detects that data
    and adds 4 conditional columns to the Transactions sheet so the
    accountant can see the per-row classification reasoning that drove
    inclusion/exclusion for 1099 aggregation. When the data is absent
    (rule-based or multi-agent engines), the workbook is byte-identical
    to v1.2 output.

    New optional kwargs accepted by generate_excel_report:
      - all_transactions:   list of Transaction objects including excluded
                            rows. When supplied, the Transactions sheet
                            iterates this list instead of `transactions`,
                            so excluded rows (payroll, fees, balance lines,
                            transfers) are visible too.
      - pdf_skill_metadata: dict from the PDF Skill agent (currently unused
                            by the workbook, accepted for forward compat).
      - **kwargs:           any other future kwargs are silently accepted.

v1.2 — Phase 2A presentation update:
    Title cells and section labels reframed to position the per-statement
    Excel as a STATEMENT-LEVEL BOOKKEEPING REVIEW deliverable. Cross-statement
    reconciliation and 1099 candidate review live in the Master Workbook /
    Consolidated Validation layer. Sheet names are intentionally preserved
    because other code paths (e.g. Excel-recovery fallback in
    agent_app.py.run_single_agent) reference them by hardcoded string.

Design principles:
    - Column headers encode ACCOUNTING logic, not just data fields
    - Review-needed rows are visually flagged (yellow highlight)
    - Output is IMPORTABLE into QuickBooks / 1099-ETC, not just readable
    - Every number is either a formula or clearly sourced
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from transaction_aggregator import VendorSummary, Transaction
from vendor_normalizer import NormalizedVendor
from vendor_classifier_1099 import EligibilityResult


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", start_color="1F3A5F")   # Navy
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
REVIEW_FILL  = PatternFill("solid", start_color="FFF4CC")   # Yellow — needs review
ELIGIBLE_FILL = PatternFill("solid", start_color="E8F5E9")  # Light green — 1099 required
MISC_FILL    = PatternFill("solid", start_color="E3F2FD")   # Light blue — 1099-MISC
EXEMPT_FILL  = PatternFill("solid", start_color="F5F5F5")   # Light grey — exempt
EXCLUDED_FILL = PatternFill("solid", start_color="F0F4F8")  # v1.3: very subtle grey — excluded rows

# ── v1.4 Phase 4D-plus — per-type fills for the Transactions "Transaction Type"
# CELL ONLY (Option A). These are NEW, SEPARATE constants — they modify no
# existing fill and are applied to nothing but the Type column in skill-mode.
# Zero impact on the web UI palette (that lives in frontend/index.html) or on any
# other sheet / the row-fill review-exclusion system. vendor_payment is
# intentionally NOT in the map (Option 1: the common row stays white / no fill).
TYPE_FILL_DEPOSIT    = PatternFill("solid", start_color="E3F2FD")  # light blue
TYPE_FILL_INTEREST   = PatternFill("solid", start_color="FEF9E7")  # light yellow
TYPE_FILL_CHECK      = PatternFill("solid", start_color="FCE4E4")  # soft red/rose
TYPE_FILL_FEE        = PatternFill("solid", start_color="FFE0B2")  # light amber
TYPE_FILL_BALANCE    = PatternFill("solid", start_color="FCE0A8")  # orange-yellow
TYPE_FILL_TRANSFER   = PatternFill("solid", start_color="EDE9FE")  # light purple
TYPE_FILL_STRUCTURAL = PatternFill("solid", start_color="F0F4F8")  # light grey

# machine transaction_type → Type-cell fill. Absent key ⇒ no fill (white).
TYPE_CELL_FILLS = {
    "deposit":         TYPE_FILL_DEPOSIT,
    "interest":        TYPE_FILL_INTEREST,
    "reimbursement":   TYPE_FILL_INTEREST,
    "check_payment":   TYPE_FILL_CHECK,
    "bank_fee":        TYPE_FILL_FEE,
    "balance_line":    TYPE_FILL_BALANCE,
    "transfer":        TYPE_FILL_TRANSFER,
    "owner_draw":      TYPE_FILL_TRANSFER,
    "payroll_deposit": TYPE_FILL_STRUCTURAL,
    "metadata":        TYPE_FILL_STRUCTURAL,
    "unknown":         TYPE_FILL_STRUCTURAL,
    # vendor_payment: intentionally absent → no fill (Option 1)
}

BODY_FONT = Font(name="Arial", size=10)
BODY_FONT_MUTED = Font(name="Arial", size=10, color="6B7280")  # v1.3: muted for excluded rows
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row_num].height = 22


def set_column_widths(ws, widths: dict[str, int]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _has_classifier_data(transactions) -> bool:
    """
    v1.3: detect whether the PDF Skill engine populated classifier fields.

    Returns True if any transaction has a non-default `transaction_type`
    OR any classifier-specific field is set. Used to decide whether to
    add the 4 conditional columns to the Transactions sheet.

    Rule-based and multi-agent engines do not set these fields, so this
    returns False for them and the workbook stays byte-identical to v1.2.
    """
    for t in transactions or []:
        # transaction_type defaults to "vendor_payment" if classifier didn't
        # set it explicitly. If we see ANY other value, classifier ran.
        ttype = getattr(t, "transaction_type", None)
        if ttype and ttype != "vendor_payment":
            return True
        # Defensive: also check the boolean fields. If False (default for
        # include_for_1099) or True (default for excluded), the classifier
        # data is present.
        if getattr(t, "exclusion_reason", None):
            return True
        if getattr(t, "review_required", False):
            return True
    return False


def _friendly_transaction_type(ttype: str) -> str:
    """Map machine transaction_type to accountant-friendly label."""
    if not ttype:
        return ""
    labels = {
        "vendor_payment":  "Vendor payment",
        "check_payment":   "Check",
        "deposit":         "Deposit",
        "payroll_deposit": "Payroll deposit",
        "balance_line":    "Balance line",
        "transfer":        "Transfer",
        "bank_fee":        "Bank fee",
        "interest":        "Interest",
        "reimbursement":   "Reimbursement",
        "owner_draw":      "Owner draw",
        "metadata":        "Metadata row",
        "unknown":         "Unclassified",
    }
    return labels.get(ttype, ttype.replace("_", " ").title())


# ---------------------------------------------------------------------------
# v1.4 Phase 4D — Summary Stats card-mirroring helpers
# ---------------------------------------------------------------------------

# Canonical type order — mirrors the frontend's TRANSACTION_TYPE_ORDER so the
# Activity Classification line on the Excel reads identically to the web card.
_TXN_TYPE_ORDER = [
    "vendor_payment", "check_payment", "deposit", "payroll_deposit",
    "balance_line", "transfer", "bank_fee", "interest",
    "reimbursement", "owner_draw", "metadata", "unknown",
]

# Plural friendly labels for count contexts (Activity Classification / excluded
# parenthetical). Mirrors the frontend's TRANSACTION_TYPE_LABELS.
_TXN_TYPE_LABELS_PLURAL = {
    "vendor_payment":  "Vendor payments",
    "check_payment":   "Checks",
    "deposit":         "Deposits",
    "payroll_deposit": "Payroll deposits",
    "balance_line":    "Balance lines",
    "transfer":        "Transfers",
    "bank_fee":        "Bank fees",
    "interest":        "Interest",
    "reimbursement":   "Reimbursements",
    "owner_draw":      "Owner draws",
    "metadata":        "Metadata rows",
    "unknown":         "Unclassified",
}


def _type_label_plural(ttype: str) -> str:
    """Plural-friendly label for a transaction type (count context)."""
    return _TXN_TYPE_LABELS_PLURAL.get(ttype, ttype.replace("_", " ").title())


def _activity_classification_phrase(breakdown: dict) -> str:
    """
    Build the Activity Classification phrase, e.g.
    "Vendor payments 7 · Deposits 1 · Bank fees 1".

    Enumerates types in canonical order (only count > 0), then any
    unenumerated types defensively. Returns "" if breakdown is empty.
    Mirrors the frontend's renderActivityClassification.
    """
    if not breakdown or not isinstance(breakdown, dict):
        return ""
    parts = []
    for t in _TXN_TYPE_ORDER:
        c = breakdown.get(t)
        if c and c > 0:
            parts.append(f"{_type_label_plural(t)} {c}")
    for t in breakdown:
        if t in _TXN_TYPE_ORDER:
            continue
        c = breakdown.get(t)
        if c and c > 0:
            parts.append(f"{t.replace('_', ' ').title()} {c}")
    return " · ".join(parts)


def _excluded_breakdown_phrase(breakdown: dict) -> str:
    """
    Build the excluded-types parenthetical, e.g. "1 deposit, 1 bank fee".

    Enumerates non-vendor_payment types with count > 0, singular/plural
    aware. Returns "" if nothing excluded. Mirrors the frontend's
    formatExcludedBreakdown (which lowercases labels for inline prose).
    """
    if not breakdown or not isinstance(breakdown, dict):
        return ""
    parts = []

    def _singularize(label_plural: str) -> str:
        # Cheap singularization for the handful of labels we emit.
        special = {
            "Checks": "check",
            "Deposits": "deposit",
            "Payroll deposits": "payroll deposit",
            "Balance lines": "balance line",
            "Transfers": "transfer",
            "Bank fees": "bank fee",
            "Reimbursements": "reimbursement",
            "Owner draws": "owner draw",
            "Metadata rows": "metadata row",
            "Interest": "interest",
            "Unclassified": "unclassified",
        }
        return special.get(label_plural, label_plural.lower())

    for t in _TXN_TYPE_ORDER:
        if t == "vendor_payment":
            continue
        c = breakdown.get(t)
        if c and c > 0:
            lbl_pl = _type_label_plural(t)
            lbl = lbl_pl.lower() if c != 1 else _singularize(lbl_pl)
            parts.append(f"{c} {lbl}")
    for t in breakdown:
        if t == "vendor_payment" or t in _TXN_TYPE_ORDER:
            continue
        c = breakdown.get(t)
        if c and c > 0:
            parts.append(f"{c} {t.replace('_', ' ')}")
    return ", ".join(parts)


# v1.4 Phase 4E — singularizer shared by the excluded-phrase builders.
def _singularize_type_label(label_plural: str) -> str:
    special = {
        "Checks": "check",
        "Deposits": "deposit",
        "Payroll deposits": "payroll deposit",
        "Balance lines": "balance line",
        "Transfers": "transfer",
        "Bank fees": "bank fee",
        "Reimbursements": "reimbursement",
        "Owner draws": "owner draw",
        "Metadata rows": "metadata row",
        "Interest": "interest",
        "Unclassified": "unclassified",
    }
    return special.get(label_plural, label_plural.lower())


def _excluded_phrase_from_counts(excluded_counts: dict) -> str:
    """
    v1.4 Phase 4E — build the excluded parenthetical from a {transaction_type:
    count} dict of rows that are ACTUALLY excluded (include_for_1099 == False).

    This is the corrected source for the Processing Details parenthetical. The
    older _excluded_breakdown_phrase(breakdown) used the full type breakdown,
    which wrongly pulled in INCLUDED non-vendor types (notably checks — a check
    to a payee is an included, potentially-1099-reportable payment, not an
    excluded row) and produced a parenthetical whose items didn't sum to the
    stated excluded count. Sourcing both the count and this phrase from the same
    include==False rows keeps them consistent by construction.

    Enumerates in canonical type order, then any unenumerated types. Singular/
    plural aware. Returns "" if nothing excluded.
    """
    if not excluded_counts or not isinstance(excluded_counts, dict):
        return ""
    parts = []
    for t in _TXN_TYPE_ORDER:
        c = excluded_counts.get(t)
        if c and c > 0:
            lbl_pl = _type_label_plural(t)
            lbl = lbl_pl.lower() if c != 1 else _singularize_type_label(lbl_pl)
            parts.append(f"{c} {lbl}")
    for t in excluded_counts:
        if t in _TXN_TYPE_ORDER:
            continue
        c = excluded_counts.get(t)
        if c and c > 0:
            parts.append(f"{c} {t.replace('_', ' ')}")
    return ", ".join(parts)


def _fmt_recon_amount(n) -> str:
    """Format a reconciliation figure as currency, e.g. -$150.00 / $3,000.00.

    Mirrors the frontend's fmtReconAmount. None → em dash.
    """
    if n is None:
        return "—"
    v = float(n)
    sign = "-" if v < 0 else ""
    return f"{sign}${abs(v):,.2f}"


# ---------------------------------------------------------------------------
# Sheet 1: Vendor Summary (statement-level bookkeeping review hero sheet)
# ---------------------------------------------------------------------------

VENDOR_SUMMARY_COLUMNS = [
    ("Vendor Name",         22, "left"),
    ("Entity Type",         12, "center"),
    ("Total Paid ($)",      15, "right"),
    ("# Payments",          11, "center"),
    ("First Payment",       13, "center"),
    ("Last Payment",        13, "center"),
    ("1099 Eligible",       14, "center"),
    ("W-9 on File",         12, "center"),
    ("Category",            14, "center"),
    ("Confidence",          11, "center"),
    ("Review Needed",       14, "center"),
    ("Review Reason",       40, "left"),
    ("Raw Name Variants",   40, "left"),
]


def write_vendor_summary_sheet(
    wb: Workbook,
    summaries: list[VendorSummary],
    eligibility: dict[str, EligibilityResult] = None,
):
    # Sheet name unchanged ("Vendor Summary") — referenced elsewhere by string.
    # v1.4 Phase 4D: position moved to index 1 (Summary Stats now leads at 0).
    ws = wb.create_sheet("Vendor Summary", 1)

    # Title row — Phase 2A: statement-level bookkeeping framing
    ws["A1"] = "STATEMENT-LEVEL BOOKKEEPING REVIEW"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    ws.merge_cells("A1:M1")

    ws["A2"] = f"Single-statement view · Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A2:M2")

    # Header row (row 4)
    for col_idx, (label, width, _) in enumerate(VENDOR_SUMMARY_COLUMNS, start=1):
        ws.cell(row=4, column=col_idx, value=label)
    style_header_row(ws, 4, len(VENDOR_SUMMARY_COLUMNS))

    # Column widths
    for col_idx, (_, width, _) in enumerate(VENDOR_SUMMARY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    row = 5
    for s in summaries:
        elig = (eligibility or {}).get(s.canonical_name)
        eligible_val = elig.form_type if elig else "TBD"
        w9_val = elig.w9_on_file if elig else "TBD"

        row_data = [
            s.canonical_name,
            s.entity_type or "Individual?",
            s.total_amount,
            s.transaction_count,
            s.first_payment_date or "",
            s.last_payment_date or "",
            eligible_val,
            w9_val,
            classify_category(s),
            s.match_confidence,
            "YES" if s.needs_review else "NO",
            "; ".join(s.review_reasons),
            "; ".join(s.raw_name_variants),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            _, _, align = VENDOR_SUMMARY_COLUMNS[col_idx - 1]
            cell.alignment = {"left": LEFT, "right": RIGHT, "center": CENTER}[align]

        # Format amount as currency
        ws.cell(row=row, column=3).number_format = '$#,##0.00;($#,##0.00);-'
        # Format confidence as percentage
        ws.cell(row=row, column=10).number_format = '0%'

        # Row color: eligibility takes priority over review flag
        if elig:
            if elig.form_type in ("1099-NEC",):
                row_fill = ELIGIBLE_FILL
            elif elig.form_type == "1099-MISC":
                row_fill = MISC_FILL
            elif elig.form_type == "REVIEW":
                row_fill = REVIEW_FILL
            else:
                row_fill = None  # EXEMPT — no fill
        elif s.needs_review:
            row_fill = REVIEW_FILL
        else:
            row_fill = None

        if row_fill:
            for col_idx in range(1, len(VENDOR_SUMMARY_COLUMNS) + 1):
                ws.cell(row=row, column=col_idx).fill = row_fill

        row += 1

    # Totals row at the bottom
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{row - 1})")
    ws.cell(row=total_row, column=3).number_format = '$#,##0.00'
    ws.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=total_row, column=4, value=f"=SUM(D5:D{row - 1})")
    ws.cell(row=total_row, column=4).font = Font(name="Arial", bold=True, size=11)
    for col_idx in range(1, len(VENDOR_SUMMARY_COLUMNS) + 1):
        ws.cell(row=total_row, column=col_idx).border = Border(
            top=Side(style="medium", color="1F3A5F")
        )

    # Freeze header row
    ws.freeze_panes = "A5"


def classify_category(summary: VendorSummary) -> str:
    """
    Simple category classification.
    A future release will replace this with a smarter classifier that reads memos.
    For now, use a heuristic based on entity type + common vendor patterns.
    """
    name_upper = summary.canonical_name.upper()

    # Known utility/corporate patterns
    utility_keywords = ["VERIZON", "COMCAST", "PG&E", "AT&T", "ELECTRIC", "GAS", "WATER"]
    retail_keywords = ["AMAZON", "HOME DEPOT", "STAPLES", "OFFICE DEPOT", "WALMART", "TARGET", "COSTCO"]

    if any(k in name_upper for k in utility_keywords):
        return "Utility"
    if any(k in name_upper for k in retail_keywords):
        return "Supplies/Retail"
    if summary.entity_type in ("LLC", "INC", "CORP"):
        return "Contractor"
    if "UNRESOLVED" in name_upper:
        return "Unclear"
    return "Unclear"


# ---------------------------------------------------------------------------
# Sheet 2: Transactions (detail view)
# ---------------------------------------------------------------------------

# v1.3: base columns (8). When PDF Skill data is present, 4 more columns
# are appended for transaction-level classification.
TXN_BASE_HEADERS = [
    ("Date", 12),
    ("Raw Description", 35),
    ("Canonical Vendor", 25),
    ("Entity Type", 12),
    ("Amount ($)", 14),
    ("Source", 12),
    ("Confidence", 12),
    ("Review?", 10),
]

TXN_CLASSIFIER_HEADERS = [
    ("Transaction Type", 18),
    ("Include for 1099", 16),
    ("Exclusion Reason", 28),
    ("Review Required", 16),
]


def write_transactions_sheet(
    wb: Workbook,
    transactions: list[Transaction],
    normalized: list[NormalizedVendor],
    all_transactions: list = None,
):
    """
    Write the Transactions sheet.

    v1.3 behavior:
      - If `all_transactions` is supplied AND it contains classifier data
        (PDF Skill engine), iterate that list so excluded rows are visible
        alongside included ones, and append the 4 classifier columns.
      - If `all_transactions` is None OR has no classifier data
        (rule-based / multi-agent engines), behavior is identical to v1.2:
        iterate `transactions` parallel with `normalized`, 8 columns only.

    The decision is data-driven, not engine-name-driven, so it's robust
    even if a future engine sets a subset of fields.
    """
    # Sheet name unchanged ("Transactions") — referenced elsewhere by string.
    # v1.4 Phase 4D: position pinned to index 2 (after Summary Stats, Vendor Summary).
    ws = wb.create_sheet("Transactions", 2)

    # Decide which iteration mode to use
    use_skill_mode = bool(all_transactions) and _has_classifier_data(all_transactions)

    # Title — engine-aware
    if use_skill_mode:
        ws["A1"] = "TRANSACTIONS — THIS STATEMENT (with row-level classification)"
    else:
        ws["A1"] = "TRANSACTIONS — THIS STATEMENT"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    # Merge title across all columns we're going to use
    total_cols = len(TXN_BASE_HEADERS) + (len(TXN_CLASSIFIER_HEADERS) if use_skill_mode else 0)
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")

    # Build header list
    headers = list(TXN_BASE_HEADERS)
    if use_skill_mode:
        headers = headers + list(TXN_CLASSIFIER_HEADERS)

    for col_idx, (label, width) in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=label)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    style_header_row(ws, 3, len(headers))

    # Build a normalized-lookup dict so we can find canonical vendor info
    # by transaction description even when iterating `all_transactions`
    # (which is the union list, not necessarily aligned with `normalized`).
    norm_by_desc = {}
    for txn, norm in zip(transactions or [], normalized or []):
        norm_by_desc[(txn.description, txn.amount)] = norm

    # Choose iteration source
    if use_skill_mode:
        # Iterate the FULL list (includes excluded). Look up normalized by
        # (description, amount) — included rows will hit; excluded rows miss
        # and get blank vendor info, which is correct.
        iter_source = [
            (t, norm_by_desc.get((t.description, t.amount)))
            for t in all_transactions
        ]
    else:
        # v1.2 behavior — zip transactions with normalized.
        iter_source = list(zip(transactions or [], normalized or []))

    row = 4
    for txn, norm in iter_source:
        # Determine row classification flags (PDF Skill case)
        is_excluded = False
        if use_skill_mode:
            include_flag = getattr(txn, "include_for_1099", True)
            is_excluded = not include_flag

        # Vendor info — present for included rows (norm exists), blank/synthetic
        # for excluded rows where norm is None.
        if norm is not None:
            canonical_vendor = norm.canonical_name
            entity_type = norm.entity_type or ""
            confidence = norm.match_confidence
            needs_review = norm.needs_review
        else:
            canonical_vendor = ""
            entity_type = ""
            confidence = None
            needs_review = False

        values = [
            txn.date or "",
            txn.description,
            canonical_vendor,
            entity_type,
            txn.amount,
            txn.source,
            confidence,
            "YES" if needs_review else "",
        ]

        if use_skill_mode:
            ttype_raw = getattr(txn, "transaction_type", "") or ""
            include_flag = getattr(txn, "include_for_1099", True)
            excl_reason = getattr(txn, "exclusion_reason", "") or ""
            review_required = getattr(txn, "review_required", False)

            values.extend([
                _friendly_transaction_type(ttype_raw),
                "YES" if include_flag else "NO",
                excl_reason,
                "YES" if review_required else "",
            ])

        # Write the row
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            # Muted font for excluded rows so accountant sees them but
            # focus stays on included rows. Skip muting for review-flagged
            # rows (they have their own visual treatment below).
            if is_excluded and not needs_review:
                cell.font = BODY_FONT_MUTED
            else:
                cell.font = BODY_FONT
            cell.border = THIN_BORDER

        # Number formats
        ws.cell(row=row, column=5).number_format = '$#,##0.00'  # Amount
        if confidence is not None:
            ws.cell(row=row, column=7).number_format = '0%'  # Confidence

        # Row fill — three tiers of priority:
        #   1. needs_review (yellow) — highest
        #   2. is_excluded (subtle grey) — medium
        #   3. no fill — included, no review needed
        if needs_review:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = REVIEW_FILL
        elif is_excluded:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = EXCLUDED_FILL

        # v1.4 Phase 4D-plus — per-type color on the "Transaction Type" CELL only
        # (Option A). Runs AFTER the row-fill block so it overrides the row fill
        # on this one cell, leaving the row's review/exclusion fill on every other
        # column. Skill-mode only (the Type column exists only then). Absent type
        # (or vendor_payment) ⇒ no override, so the cell keeps whatever the row
        # fill gave it (white for a normal included vendor payment).
        if use_skill_mode:
            # Transaction Type is the first classifier column = column
            # len(TXN_BASE_HEADERS) + 1.
            type_col = len(TXN_BASE_HEADERS) + 1
            type_fill = TYPE_CELL_FILLS.get(ttype_raw)
            if type_fill is not None:
                ws.cell(row=row, column=type_col).fill = type_fill

        row += 1

    ws.freeze_panes = "A4"

    # v1.4 Phase 4D-plus — native Excel AutoFilter on the header row (Option X).
    # Gives interactive sort/filter dropdowns on every column; the user reorders
    # in Excel by date / vendor / type / amount / any column, live and reversible.
    # Header is row 3; data starts at row 4. Guard against zero data rows.
    last_data_row = row - 1
    if last_data_row >= 4:
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A3:{last_col_letter}{last_data_row}"


# ---------------------------------------------------------------------------
# Sheet 3: Summary Statistics
# ---------------------------------------------------------------------------

# v1.4 Phase 4D — reconciliation verdict/diff font colors. Mirrors the web
# card's MODERATE needs_review treatment (amber text + ⚠, no loud band) and
# its balanced treatment (green check). Font color, not fill, keeps it calm.
RECON_OK_FONT    = Font(name="Arial", bold=True, size=11, color="16A34A")   # green (--ok)
RECON_WARN_FONT  = Font(name="Arial", bold=True, size=11, color="EA580C")   # amber (--warn)
RECON_NOTE_FONT  = Font(name="Arial", italic=True, size=9, color="6B7280")  # muted note
SECTION_FONT     = Font(name="Arial", bold=True, size=11, color="1F3A5F")   # navy section head


def write_summary_stats_sheet(
    wb: Workbook,
    summaries: list[VendorSummary],
    transactions: list[Transaction],
    *,
    reconciliation_snapshot: dict = None,
    breakdown: dict = None,
    all_transactions: list = None,
    confidence: float = None,
    extraction_check: dict = None,
):
    """
    v1.4 Phase 4D — the Summary Stats sheet is now the workbook's LANDING sheet
    (index 0) and mirrors the web Per-Statement card's expansion, top to bottom:

        1. Title + timestamp
        2. Statement Processing Details   (rows identified / included / excluded)
        3. Activity Classification        (type breakdown — PDF Skill only)
        4. Vendor / 1099 Review           (totals line — both engines)
        5. Statement Reconciliation       (the waterfall — PDF Skill w/ snapshot)
        6. Bookkeeping Review Signals     (review needed / over $600 — both)
        7. Scope notes                    (retained)

    Blocks 3 and 5 auto-omit gracefully when their data is absent (rule-based:
    no breakdown, no snapshot), so the sheet degrades to extraction + review
    content. Sheet TAB NAME stays "Summary Stats" (load-bearing); only the
    position changed (to index 0, set by create_sheet below).
    """
    # Sheet name unchanged ("Summary Stats") — referenced elsewhere by string.
    # v1.4 Phase 4D: position pinned to index 0 (this is now the landing sheet).
    ws = wb.create_sheet("Summary Stats", 0)

    # Title cell text unchanged (non-load-bearing visual heading; user's choice).
    ws["A1"] = "STATEMENT BOOKKEEPING SUMMARY"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
    ws.merge_cells("A1:C1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A2:C2")

    # ── Derived stats (metadata, computed here — not model values) ──
    total_txns = len(transactions)
    total_vendors = len(summaries)
    total_amount = sum(s.total_amount for s in summaries)
    vendors_over_600 = sum(1 for s in summaries if s.total_amount >= 600)
    review_needed = sum(1 for s in summaries if s.needs_review)
    llc_vendors = sum(1 for s in summaries if s.entity_type in ("LLC", "INC", "CORP"))

    # Statement-level confidence. Use the value the PIPELINE computed and the
    # web card displays (response dict "confidence" = avg row-level confidence,
    # a 0-1 float) so the Excel reads identically to the card. Only fall back to
    # a vendor-summary estimate if the caller didn't pass it (e.g. older callers).
    if confidence is not None:
        confidence_pct = round(float(confidence) * 100)
    elif summaries:
        # Defensive fallback only — avg of vendor match_confidence (0-1).
        vals = [(getattr(s, "match_confidence", 0) or 0) for s in summaries]
        confidence_pct = round((sum(vals) / len(vals)) * 100) if vals else 0
    else:
        confidence_pct = 0

    # Included vs excluded counts for Processing Details. Sourced from the actual
    # rows in `all_transactions` (each carries include_for_1099). v1.4 Phase 4E:
    # we also tally the excluded rows BY transaction_type here, so the parenthetical
    # and the count come from the SAME include==False rows — keeping them consistent
    # and (critically) keeping INCLUDED types like checks out of the excluded list.
    included_count = total_txns
    excluded_count = 0
    excluded_by_type = {}
    if all_transactions:
        for t in all_transactions:
            if not getattr(t, "include_for_1099", True):
                excluded_count += 1
                ttype = getattr(t, "transaction_type", "") or "unknown"
                excluded_by_type[ttype] = excluded_by_type.get(ttype, 0) + 1
        included_count = len(all_transactions) - excluded_count
    parsed_count = included_count + excluded_count

    has_breakdown = bool(breakdown and isinstance(breakdown, dict)
                         and any(v for v in breakdown.values()))

    row = 4

    def _section(title: str):
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = SECTION_FONT
        row += 1

    def _line(text: str, *, font=None, merge_to=3):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = font or BODY_FONT
        if merge_to and merge_to > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=merge_to)
        row += 1

    def _blank():
        nonlocal row
        row += 1

    # ── Block 2 — Statement Processing Details ──
    _section("STATEMENT PROCESSING DETAILS")
    if has_breakdown:
        if excluded_count == 0:
            _line(f"{parsed_count} row{'' if parsed_count == 1 else 's'} identified. "
                  f"All {included_count} included as vendor payments for 1099 aggregation.")
        else:
            # v1.4 Phase 4E — parenthetical built from the SAME include==False
            # rows the count uses (excluded_by_type), not the full breakdown dict.
            # This keeps the count and the list consistent and excludes INCLUDED
            # types like checks.
            excl_phrase = _excluded_phrase_from_counts(excluded_by_type)
            detail = (f"{parsed_count} rows identified. "
                      f"{included_count} included as vendor payments for 1099 aggregation. "
                      f"{excluded_count} excluded")
            detail += f" ({excl_phrase})." if excl_phrase else "."
            _line(detail)
    else:
        # Rule-based / no breakdown — generic line, mirrors the card fallback.
        _line(f"{total_txns} transaction{'' if total_txns == 1 else 's'} processed. "
              f"No row-level classification available for this engine.")
    _blank()

    # ── Block 3 — Activity Classification (PDF Skill only) ──
    if has_breakdown:
        phrase = _activity_classification_phrase(breakdown)
        if phrase:
            _section("ACTIVITY CLASSIFICATION")
            _line(phrase)
            _blank()

    # ── Block 4 — Vendor / 1099 Review (both engines) ──
    _section("VENDOR / 1099 REVIEW")
    _line(f"Included Total {_fmt_recon_amount(total_amount)} · "
          f"Vendors {total_vendors} · "
          f"Review Needed {review_needed} · "
          f"Over $600 {vendors_over_600} · "
          f"Confidence {confidence_pct}%")
    _blank()

    # ── Block 5 — Statement Reconciliation (PDF Skill w/ complete snapshot) ──
    snap = reconciliation_snapshot
    recon_available = bool(
        snap and isinstance(snap, dict)
        and snap.get("status") in ("balanced", "needs_review")
        and snap.get("extraction_complete")
    )
    if recon_available:
        _section("STATEMENT RECONCILIATION")
        # Waterfall rows: (operator, label, value)
        waterfall = [
            ("",  "Beginning balance",            snap.get("beginning_balance")),
            ("+", "Deposits & credits",           snap.get("total_deposits")),
            ("−", "Withdrawals",                  snap.get("total_withdrawals")),
            ("−", "Checks",                       snap.get("checks")),
            ("−", "Transfers",                    snap.get("transfers")),
            ("−", "Fees & charges",               snap.get("fees")),
        ]
        for op, label, val in waterfall:
            ws.cell(row=row, column=1, value=op).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value=label).font = BODY_FONT
            amt = ws.cell(row=row, column=3, value=val if val is not None else 0.0)
            amt.font = BODY_FONT
            amt.number_format = '$#,##0.00'
            amt.alignment = RIGHT
            row += 1

        # Calculated ending (bold, thin top border to match the card's rule line)
        ws.cell(row=row, column=1, value="=").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value="Calculated ending").font = Font(name="Arial", bold=True, size=10)
        calc = ws.cell(row=row, column=3, value=snap.get("calculated_ending_balance") or 0.0)
        calc.font = Font(name="Arial", bold=True, size=10)
        calc.number_format = '$#,##0.00'
        calc.alignment = RIGHT
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="888888"))
        row += 1

        # Reported ending (as stated)
        ws.cell(row=row, column=2, value="Reported ending (as stated)").font = BODY_FONT
        rep = ws.cell(row=row, column=3, value=snap.get("reported_ending_balance") or 0.0)
        rep.font = BODY_FONT
        rep.number_format = '$#,##0.00'
        rep.alignment = RIGHT
        row += 1

        # Difference (amber when needs_review, muted when balanced; thin top border)
        is_balanced = snap.get("status") == "balanced"
        ws.cell(row=row, column=2, value="Difference").font = Font(name="Arial", bold=True, size=10)
        diff = ws.cell(row=row, column=3, value=snap.get("difference") or 0.0)
        diff.number_format = '$#,##0.00'
        diff.alignment = RIGHT
        diff.font = (Font(name="Arial", bold=True, size=10, color="6B7280") if is_balanced
                     else Font(name="Arial", bold=True, size=10, color="EA580C"))
        for col in (2, 3):
            ws.cell(row=row, column=col).border = Border(top=Side(style="thin", color="CCCCCC"))
        row += 1

        # Verdict line
        if is_balanced:
            ws.cell(row=row, column=1, value="✓").font = RECON_OK_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value="Balanced").font = RECON_OK_FONT
        else:
            ws.cell(row=row, column=1, value="⚠").font = RECON_WARN_FONT
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value="Needs Review").font = RECON_WARN_FONT
        row += 1

        # Verbatim model note (italic, muted) — mirrors the card's .ps-recon-notes
        note = snap.get("notes")
        if note:
            c = ws.cell(row=row, column=1, value=str(note))
            c.font = RECON_NOTE_FONT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 30
            row += 1
        _blank()

    # ── Block 5b — Extraction Cross-Check (Source B, PDF Skill only) ──
    # Companion to the reconciliation waterfall above. Where Block 5 shows
    # whether the statement's stated math balances, Block 5b shows whether
    # our extracted rows sum to those stated activity totals — catching
    # missed/miscounted rows during extraction. Status: complete (all 5
    # buckets match within tolerance) / incomplete (one or more deltas
    # exceed tolerance) / unavailable (no snapshot to compare against).
    # Omitted entirely when extraction_check is absent or unavailable.
    ec = extraction_check
    ec_available = bool(
        ec and isinstance(ec, dict)
        and ec.get("status") in ("complete", "incomplete")
    )
    if ec_available:
        _section("EXTRACTION CROSS-CHECK")
        is_complete = ec.get("status") == "complete"

        # Header line with verdict — green ✓ when complete, amber ⚠ when not.
        verdict_label = "Complete — extracted rows match stated totals" if is_complete \
                        else "Incomplete — row sums diverge from stated totals"
        verdict_icon  = "✓" if is_complete else "⚠"
        verdict_font  = RECON_OK_FONT if is_complete else RECON_WARN_FONT
        ws.cell(row=row, column=1, value=verdict_icon).font = verdict_font
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=verdict_label).font = verdict_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

        # Column header row for the bucket table
        hdr_row = row
        for col, label in enumerate(["Bucket", "Stated", "Row Sum", "Delta"], start=1):
            c = ws.cell(row=hdr_row, column=col, value=label)
            c.font = Font(name="Arial", bold=True, size=9, color="475569")
            c.alignment = CENTER if col != 1 else LEFT
        # We have 3 default cols (A,B,C) but need 4 — extend D for this block.
        # The column widths block at the end of this function sets A=50,B=28,C=16;
        # for this block we use narrower columns. Override D width specifically.
        ws.column_dimensions["D"].width = 14
        row += 1

        # Five bucket rows
        buckets = [
            ("Deposits",    "deposits"),
            ("Withdrawals", "withdrawals"),
            ("Checks",      "checks"),
            ("Transfers",   "transfers"),
            ("Fees",        "fees"),
        ]
        for label, key in buckets:
            stated  = ec.get(f"{key}_stated")
            row_sum = ec.get(f"{key}_row_sum")
            delta   = ec.get(f"{key}_delta")

            # Skip rows where everything's None/zero — keeps the table compact
            # for statements that genuinely don't have checks/transfers.
            if stated is None and (row_sum is None or abs(row_sum) < 0.01):
                continue

            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT

            st_cell = ws.cell(row=row, column=2,
                              value=stated if stated is not None else "—")
            st_cell.font = BODY_FONT
            st_cell.alignment = RIGHT
            if stated is not None:
                st_cell.number_format = '$#,##0.00'

            rs_cell = ws.cell(row=row, column=3,
                              value=row_sum if row_sum is not None else 0.0)
            rs_cell.font = BODY_FONT
            rs_cell.alignment = RIGHT
            rs_cell.number_format = '$#,##0.00'

            # Delta: amber font when non-zero (over tolerance), muted otherwise.
            dl_cell = ws.cell(row=row, column=4,
                              value=delta if delta is not None else 0.0)
            dl_cell.number_format = '$#,##0.00;-$#,##0.00;-'
            dl_cell.alignment = RIGHT
            if delta is not None and abs(delta) > 0.01:
                dl_cell.font = Font(name="Arial", bold=True, size=10, color="EA580C")
            else:
                dl_cell.font = Font(name="Arial", size=10, color="6B7280")
            row += 1
        _blank()

    # ── Block 6 — Bookkeeping Review Signals (both engines) ──
    signal_parts = []
    if review_needed > 0:
        signal_parts.append(f"Review needed {review_needed}")
    if vendors_over_600 > 0:
        signal_parts.append(f"Over $600 {vendors_over_600}")
    if signal_parts:
        _section("BOOKKEEPING REVIEW SIGNALS")
        _line(" · ".join(signal_parts))
        _blank()

    # Column widths — A wide for labels/waterfall, B for waterfall labels,
    # C for the right-aligned currency column.
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16

    # ── Block 7 — Scope notes (retained) ──
    row += 1
    ws.cell(row=row, column=1, value="SCOPE NOTES").font = SECTION_FONT
    row += 1
    notes = [
        "Per-statement output supports statement-level bookkeeping review.",
        "Cross-statement vendor comparison and 1099 threshold review are handled",
        "in the Master Workbook / Consolidated Validation layer.",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note).font = Font(name="Arial", italic=True, size=9, color="666666")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_excel_report(
    output_path: str,
    transactions: list[Transaction],
    normalized: list[NormalizedVendor],
    summaries: list[VendorSummary],
    eligibility: dict[str, EligibilityResult] = None,
    *,
    all_transactions: list = None,
    pdf_skill_metadata: dict = None,
    reconciliation_snapshot: dict = None,   # v1.4 Phase 4D
    breakdown: dict = None,                  # v1.4 Phase 4D
    confidence: float = None,                # v1.4 Phase 4D — statement-level
    extraction_check: dict = None,           # v1.4 Phase 4 — Source B
    **kwargs,
):
    """Generate the full three-sheet per-statement Excel report.

    v1.4 Phase 4D — workbook restructure + reconciliation:
      - Sheet order changes so "Summary Stats" opens FIRST (index 0), acting
        as a one-page statement overview that mirrors the web Per-Statement
        card. "Vendor Summary" (1) and "Transactions" (2) follow as detail
        sheets. Sheet TAB NAMES are unchanged — only positions change — so
        hardcoded-string lookups (agent_app.py.run_single_agent) still resolve.
      - Summary Stats gains five card-mirroring blocks: Statement Processing
        Details, Activity Classification, Vendor / 1099 Review, Statement
        Reconciliation (the waterfall), and Bookkeeping Review Signals.
      - `reconciliation_snapshot` (computed server-side in pipeline) drives the
        reconciliation waterfall; `breakdown` (type→count dict) drives Activity
        Classification and the excluded parenthetical. Both are optional: when
        absent (rule-based engine), the dependent blocks are omitted gracefully
        and the sheet degrades to its prior extraction/review content.

    v1.3 — accepts optional kwargs `all_transactions` and `pdf_skill_metadata`
    from the PDF Skill pipeline path. When `all_transactions` is supplied and
    contains classifier data, the Transactions sheet renders 4 additional
    columns and shows excluded rows. When absent, output is identical to v1.2.

    The trailing `**kwargs` swallows any future params without crashing so
    callers can evolve independently of this generator.
    """
    # `pdf_skill_metadata` is currently unused but accepted for forward
    # compatibility — future versions could surface PDF Skill cost/tool-call
    # info on Summary Stats sheet.
    _ = pdf_skill_metadata
    _ = kwargs

    wb = Workbook()
    wb.remove(wb.active)

    # v1.4 Phase 4D — create in display order so Summary Stats opens first.
    # Each writer passes its own explicit sheet index (see write_* functions).
    write_summary_stats_sheet(wb, summaries, transactions,
                              reconciliation_snapshot=reconciliation_snapshot,
                              breakdown=breakdown,
                              all_transactions=all_transactions,
                              confidence=confidence,
                              extraction_check=extraction_check)
    write_vendor_summary_sheet(wb, summaries, eligibility)
    write_transactions_sheet(wb, transactions, normalized,
                             all_transactions=all_transactions)

    wb.save(output_path)
    return output_path
