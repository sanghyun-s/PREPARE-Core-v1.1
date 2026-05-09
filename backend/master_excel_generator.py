"""
Master Excel Generator — v1.1
------------------------------
Produces a single consolidated workbook from multiple agent outputs.

v1.1 changes from v1.0:
    * NEW: Executive Summary sheet at index 0 (workbook opens here by default).
      Contents: KPI grid, validation overview counts, top 10 vendors by total
      amount, run metadata. Mirrors the dashboard's Workspace view.
    * Currency rounding helper to prevent floating-point leaks
      (e.g., 87.43000000000001 → 87.43).
    * Master Vendor Summary now freezes column A (vendor name) so it stays
      visible when scrolling right through the 13-column row.
    * No changes to detection / aggregation / validation logic.

v1.0 changes (preserved):
    * AI Validation Narrative block removed
    * filename_map for original-filename rendering in user-facing cells

Sheet order (sheet count = 5):
  0. Executive Summary    ★ new in v1.1
  1. Master Vendor Summary
  2. Validation Report
  3. All Transactions
  4. Per-Agent Summary

Visual conventions (matching v0.5.1 per-agent Excel):
  • Navy header bars, ivory body
  • Yellow fill: review needed
  • Green fill: 1099-NEC required
  • Blue fill: 1099-MISC required
  • Light grey: exempt
  • Red border-left: failed/excluded files
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from validation_engine import DeterministicValidation
from review_flag_engine import ReviewFlags


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FILL    = PatternFill("solid", start_color="1F3A5F")    # Navy
HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT     = Font(name="Arial", bold=True, color="1F3A5F", size=14)
SUBTITLE_FONT  = Font(name="Arial", bold=True, color="1F3A5F", size=12)
BODY_FONT      = Font(name="Arial", size=10)
BODY_BOLD      = Font(name="Arial", bold=True, size=10)
SUBTLE_FONT    = Font(name="Arial", size=9, color="6B7280", italic=True)
KPI_LABEL_FONT = Font(name="Arial", bold=True, color="475569", size=10)
KPI_VALUE_FONT = Font(name="Arial", bold=True, color="0F172A", size=16)

REVIEW_FILL    = PatternFill("solid", start_color="FFF4CC")
ELIGIBLE_FILL  = PatternFill("solid", start_color="E8F5E9")
MISC_FILL      = PatternFill("solid", start_color="E3F2FD")
FAILED_FILL    = PatternFill("solid", start_color="FFE0E0")
KPI_FILL       = PatternFill("solid", start_color="F8FAFC")    # Soft slate

LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN     = Side(border_style="thin", color="D1D5DB")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# v1.1 helpers
# ---------------------------------------------------------------------------

def _round_currency(value) -> float:
    """
    Round a numeric value to 2 decimal places before writing to a cell.
    Prevents floating-point artifacts like 87.43000000000001 from leaking
    into the workbook. Returns 0.0 for None / non-numeric values.
    """
    if value is None:
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def _make_resolver(filename_map: dict[str, str] | None):
    """
    Return a `_resolve(label)` function that converts internal uuid-based
    statement labels into original filenames. Same as v1.0.
    """
    if not filename_map:
        return lambda label: label

    def resolve(label: str) -> str:
        if not label or label == "Combined":
            return label
        if label.endswith(".pdf"):
            stem = label[:-4]
        else:
            stem = label
        return filename_map.get(stem, label)

    return resolve


# ---------------------------------------------------------------------------
# Sheet 0 — Executive Summary (NEW in v1.1)
# ---------------------------------------------------------------------------

def write_executive_summary(
    wb: Workbook,
    agent_outputs: list[dict],
    flags_by_statement: dict[str, dict[str, ReviewFlags]],
    eligibility_by_statement: dict[str, dict],
    validation: DeterministicValidation,
    filename_map: dict[str, str] | None = None,
):
    """
    Top-of-workbook executive overview. Designed to be the first thing an
    accountant sees when they open the file.

    Layout (10 columns wide, A through J — five KPI tiles, each merged
    across two columns to give large currency values like "$28,270.94"
    enough room to render at 16pt bold without clipping):
      Row 1     — Title
      Row 2     — Subtitle (timestamp)
      Row 4     — KPI section header
      Row 5     — KPI labels    (A:B, C:D, E:F, G:H, I:J merged)
      Row 6     — KPI values    (same merges, 16pt bold)
      Row 7     — KPI sublabels (same merges)
      Row 9     — Validation Overview header
      Row 10-13 — 4 validation count rows
      Row 15    — Top Vendors header
      Row 16+   — Top 10 vendors by total amount across all statements
      Row N+2   — Run metadata footer
    """
    ws = wb.create_sheet("Executive Summary", 0)
    resolve = _make_resolver(filename_map)

    # Column widths — 10 columns. Each KPI tile occupies 2 merged columns
    # totaling 34 width-units, comfortably fitting "$28,270.94" at 16pt.
    # The Top Vendors table below also uses the wider A column for vendor
    # names which often exceed 20 chars.
    widths = [22, 12, 17, 17, 17, 17, 17, 17, 17, 17]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ──
    ws["A1"] = "EXECUTIVE SUMMARY — 1099 Pre-Reconciliation"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
    ws["A2"].font = SUBTLE_FONT
    ws.merge_cells("A2:J2")

    # ── Compute KPIs ──
    successful = [o for o in agent_outputs if o.get("status") in ("success", "partial")]
    total_transactions = sum(len(o.get("transactions", [])) for o in successful)
    unique_vendor_names = set()
    total_amount = 0.0
    vendors_over_threshold = 0
    for o in successful:
        for v in o.get("vendors", []):
            unique_vendor_names.add(v["canonical_name"])
            total_amount += v.get("total_amount", 0.0)
            if v.get("total_amount", 0) >= 600:
                vendors_over_threshold += 1
    review_needed = sum(
        sum(1 for f in flags.values() if f.needs_review)
        for flags in flags_by_statement.values()
    )

    # ── KPI tiles ──
    ws["A4"] = "KEY METRICS"
    ws["A4"].font = SUBTITLE_FONT
    ws.merge_cells("A4:J4")
    ws.row_dimensions[4].height = 20

    # Each KPI occupies 2 merged columns → 5 tiles × 2 cols = 10 cols total.
    # tile_ranges: (label_range, value_range, sublabel_range, start_col, end_col)
    tile_ranges = [
        ("A5:B5", "A6:B6", "A7:B7", 1, 2),
        ("C5:D5", "C6:D6", "C7:D7", 3, 4),
        ("E5:F5", "E6:F6", "E7:F7", 5, 6),
        ("G5:H5", "G6:H6", "G7:H7", 7, 8),
        ("I5:J5", "I6:J6", "I7:J7", 9, 10),
    ]
    kpis = [
        ("Total Transactions",   f"{total_transactions:,}",        "Across all statements"),
        ("Unique Vendors",       f"{len(unique_vendor_names):,}",  "Unique across all statements"),
        ("Total Reconciled",     f"${_round_currency(total_amount):,.2f}", "Total amount matched"),
        ("Over $600 (1099)",     f"{vendors_over_threshold:,}",    "Potential 1099 candidates"),
        ("Review Needed",        f"{review_needed:,}",             "Require human review"),
    ]
    for (label_rng, value_rng, sub_rng, start_col, end_col), (label, value, sublabel) in zip(tile_ranges, kpis):
        # Apply merges
        ws.merge_cells(label_rng)
        ws.merge_cells(value_rng)
        ws.merge_cells(sub_rng)
        # Label cell (top-left of merged range)
        c = ws.cell(row=5, column=start_col, value=label)
        c.font = KPI_LABEL_FONT
        c.alignment = CENTER
        c.fill = KPI_FILL
        # Value cell (16pt bold)
        c = ws.cell(row=6, column=start_col, value=value)
        c.font = KPI_VALUE_FONT
        c.alignment = CENTER
        c.fill = KPI_FILL
        # Sublabel cell
        c = ws.cell(row=7, column=start_col, value=sublabel)
        c.font = SUBTLE_FONT
        c.alignment = CENTER
        c.fill = KPI_FILL
        # Border around the entire merged tile (all 6 cells: 2 cols × 3 rows)
        for r in (5, 6, 7):
            for col_idx in range(start_col, end_col + 1):
                ws.cell(row=r, column=col_idx).border = BORDER
                # Fill applies to every cell in the merge for consistent rendering
                if r == 5:
                    ws.cell(row=r, column=col_idx).fill = KPI_FILL
                elif r == 6:
                    ws.cell(row=r, column=col_idx).fill = KPI_FILL
                else:
                    ws.cell(row=r, column=col_idx).fill = KPI_FILL
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 18

    # ── Validation Overview ──
    ws["A9"] = "VALIDATION OVERVIEW"
    ws["A9"].font = SUBTITLE_FONT
    ws.merge_cells("A9:J9")
    ws.row_dimensions[9].height = 20

    val_rows = [
        ("Cross-Statement Matches", len(validation.cross_matches),
         "Vendors appearing in 2 or more statements"),
        ("Name Variant Flags", len(validation.name_variants),
         "Potential duplicate vendors with different names"),
        ("Discrepancy Alerts", len(validation.amount_mismatches),
         "Possible extraction or amount mismatches"),
        ("Near-Threshold Vendors", len(validation.near_threshold),
         "Vendors close to the $600 1099 threshold"),
    ]
    row = 10
    for label, count, desc in val_rows:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=count).font = BODY_BOLD
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=desc).font = SUBTLE_FONT
        ws.cell(row=row, column=3).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 22
        # Highlight non-zero counts amber across all 10 columns
        if count > 0:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = REVIEW_FILL
        row += 1

    # ── Top Vendors ──
    row += 1
    ws.cell(row=row, column=1, value="TOP VENDORS BY PAYMENT AMOUNT").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 20
    row += 1

    # Build a global vendor totals map (sum across statements for top-N display)
    vendor_totals: dict[str, dict] = {}
    for o in successful:
        for v in o.get("vendors", []):
            name = v["canonical_name"]
            if name not in vendor_totals:
                vendor_totals[name] = {
                    "total_amount": 0.0,
                    "transaction_count": 0,
                    "entity_type": v.get("entity_type"),
                    "match_confidence": v.get("match_confidence", 1.0),
                    "needs_review": False,
                    "form_type": "TBD",
                }
            vendor_totals[name]["total_amount"] += v.get("total_amount", 0.0)
            vendor_totals[name]["transaction_count"] += v.get("transaction_count", 0)
            if v.get("entity_type") and not vendor_totals[name]["entity_type"]:
                vendor_totals[name]["entity_type"] = v.get("entity_type")

    # Roll up review flag and 1099 form type from per-statement data
    for stmt_label, flags in flags_by_statement.items():
        for vname, flag in flags.items():
            if vname in vendor_totals and flag.needs_review:
                vendor_totals[vname]["needs_review"] = True
    for stmt_label, eligs in eligibility_by_statement.items():
        for vname, e in eligs.items():
            if vname in vendor_totals and e:
                # Prefer non-EXEMPT/non-TBD form type if multiple statements give different types
                current = vendor_totals[vname]["form_type"]
                if current in ("TBD", "EXEMPT") and e.form_type not in ("TBD",):
                    vendor_totals[vname]["form_type"] = e.form_type

    top_vendors = sorted(
        vendor_totals.items(),
        key=lambda x: x[1]["total_amount"],
        reverse=True,
    )[:10]

    # Top Vendors table — 7 cols wrapped into the 10-col grid via merged cells
    # for the Vendor name column to give it room for long vendor names.
    headers = ["Vendor", "Entity Type", "Total Paid ($)", "# Payments",
               "1099 Eligible", "Review Needed", "Confidence"]
    # Vendor (1-3) | Entity Type (4) | Total Paid (5-6) | # Payments (7) | 1099 (8) | Review (9) | Conf (10)
    header_layout = [
        (1, 3, "Vendor"),
        (4, 4, "Entity Type"),
        (5, 6, "Total Paid ($)"),
        (7, 7, "# Payments"),
        (8, 8, "1099 Eligible"),
        (9, 9, "Review Needed"),
        (10, 10, "Confidence"),
    ]
    for start_col, end_col, h in header_layout:
        c = ws.cell(row=row, column=start_col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        if start_col != end_col:
            ws.merge_cells(start_row=row, start_column=start_col,
                           end_row=row, end_column=end_col)
            for col_idx in range(start_col, end_col + 1):
                ws.cell(row=row, column=col_idx).fill = HEADER_FILL
                ws.cell(row=row, column=col_idx).border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    for name, info in top_vendors:
        # Vendor name (merged cols 1-3)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        # Entity Type (col 4)
        ws.cell(row=row, column=4, value=info["entity_type"] or "Individual?").alignment = CENTER
        ws.cell(row=row, column=4).font = BODY_FONT
        # Total Paid (merged cols 5-6)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=_round_currency(info["total_amount"]))
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = RIGHT
        # # Payments (col 7)
        ws.cell(row=row, column=7, value=info["transaction_count"]).alignment = CENTER
        ws.cell(row=row, column=7).font = BODY_FONT
        # 1099 Eligible (col 8)
        ws.cell(row=row, column=8, value=info["form_type"]).alignment = CENTER
        ws.cell(row=row, column=8).font = BODY_FONT
        # Review Needed (col 9)
        ws.cell(row=row, column=9, value="YES" if info["needs_review"] else "NO").alignment = CENTER
        ws.cell(row=row, column=9).font = BODY_FONT
        # Confidence (col 10)
        ws.cell(row=row, column=10, value=info["match_confidence"])
        ws.cell(row=row, column=10).number_format = '0%'
        ws.cell(row=row, column=10).alignment = CENTER
        ws.cell(row=row, column=10).font = BODY_FONT

        # Borders across all 10 columns
        for c in range(1, 11):
            ws.cell(row=row, column=c).border = BORDER

        # Row coloring: 1099-eligible gets soft green; review-needed gets amber
        if info["form_type"] == "1099-NEC":
            row_fill = ELIGIBLE_FILL
        elif info["form_type"] == "1099-MISC":
            row_fill = MISC_FILL
        elif info["needs_review"]:
            row_fill = REVIEW_FILL
        else:
            row_fill = None
        if row_fill:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = row_fill
        row += 1

    # ── Run metadata footer ──
    row += 2
    ws.cell(row=row, column=1, value="RUN METADATA").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 20
    row += 1

    statement_names = [resolve(o["statement_label"]) for o in agent_outputs]
    successful_count = sum(1 for o in agent_outputs if o.get("status") in ("success", "partial"))
    failed_count = len(agent_outputs) - successful_count

    metadata = [
        ("Statements processed", f"{len(agent_outputs)} ({successful_count} successful, {failed_count} failed)"),
        ("Source files", ", ".join(statement_names) if statement_names else "—"),
        ("Sheets in this workbook", "5 (Executive Summary, Master Vendor Summary, Validation Report, All Transactions, Per-Agent Summary)"),
    ]
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 22
        row += 1

    # Page setup: 10 columns is wider than a portrait page can hold cleanly,
    # so set landscape orientation and fit-to-width = 1 page. Excel and
    # LibreOffice both honor these for print preview, even though screen
    # rendering (where most accountants will view this) ignores them.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


# ---------------------------------------------------------------------------
# Sheet 1 — Master Vendor Summary
# ---------------------------------------------------------------------------

MASTER_VENDOR_COLUMNS = [
    ("Vendor",                30, "left"),
    ("Source Statement",      40, "left"),
    ("Entity Type",           12, "center"),
    ("Total Paid ($)",        14, "right"),
    ("# Payments",            10, "center"),
    ("First Payment",         12, "center"),
    ("Last Payment",          12, "center"),
    ("1099 Eligible",         11, "center"),
    ("Match Confidence",      14, "center"),
    ("Extraction Confidence", 16, "center"),
    ("Review Needed",         11, "center"),
    ("Review Reasons",        55, "left"),
    ("Cross-Reference",       45, "left"),
]


def write_master_vendor_summary(
    wb: Workbook,
    agent_outputs: list[dict],
    flags_by_statement: dict[str, dict[str, ReviewFlags]],
    eligibility_by_statement: dict[str, dict],
    validation: DeterministicValidation,
    filename_map: dict[str, str] | None = None,
):
    ws = wb.create_sheet("Master Vendor Summary")
    resolve = _make_resolver(filename_map)

    ws["A1"] = "MASTER VENDOR SUMMARY — All Statements Combined"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(MASTER_VENDOR_COLUMNS))
    ws.row_dimensions[1].height = 22

    ws["A2"] = (
        f"{len(agent_outputs)} statement(s) processed. "
        "Same vendor across multiple statements appears as separate rows "
        "with cross-reference annotations."
    )
    ws["A2"].font = SUBTLE_FONT
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(MASTER_VENDOR_COLUMNS))

    for col_idx, (header, width, _) in enumerate(MASTER_VENDOR_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 30

    # Build cross-reference lookup
    name_to_statements: dict[str, list[tuple[str, float]]] = {}
    for out in agent_outputs:
        if out.get("status") not in ("success", "partial"):
            continue
        for v in out.get("vendors", []):
            name_to_statements.setdefault(v["canonical_name"], []).append(
                (out["statement_label"], v["total_amount"])
            )

    row = 5
    for out in agent_outputs:
        if out.get("status") not in ("success", "partial"):
            continue
        statement_label = out["statement_label"]
        flags = flags_by_statement.get(statement_label, {})
        elig = eligibility_by_statement.get(statement_label, {})

        for v in out.get("vendors", []):
            name = v["canonical_name"]
            f    = flags.get(name)
            e    = elig.get(name)

            xrefs = [
                f"{resolve(s)}: ${_round_currency(a):,.2f}"
                for s, a in name_to_statements.get(name, [])
                if s != statement_label
            ]
            xref_text = "; ".join(xrefs) if xrefs else ""

            row_data = [
                name,
                resolve(statement_label),
                v.get("entity_type") or "Individual?",
                _round_currency(v["total_amount"]),
                v["transaction_count"],
                v.get("first_payment_date") or "",
                v.get("last_payment_date") or "",
                e.form_type if e else "TBD",
                v.get("match_confidence", 1.0),
                f.extraction_confidence if f else 1.0,
                "YES" if (f and f.needs_review) else "NO",
                "; ".join(f.reasons) if f and f.reasons else "",
                xref_text,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.border = BORDER
                _, _, align = MASTER_VENDOR_COLUMNS[col_idx - 1]
                cell.alignment = {"left": LEFT, "right": RIGHT, "center": CENTER}[align]

            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=9).number_format = '0%'
            ws.cell(row=row, column=10).number_format = '0%'

            row_fill = None
            if e:
                if e.form_type == "1099-NEC":
                    row_fill = ELIGIBLE_FILL
                elif e.form_type == "1099-MISC":
                    row_fill = MISC_FILL
                elif e.form_type == "REVIEW":
                    row_fill = REVIEW_FILL
            if not row_fill and f and f.needs_review:
                row_fill = REVIEW_FILL

            if row_fill:
                for col_idx in range(1, len(MASTER_VENDOR_COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = row_fill

            reasons_text = row_data[11] or ""
            xref         = row_data[12] or ""
            longest = max(len(reasons_text), len(xref))
            if longest > 50:
                ws.row_dimensions[row].height = 45
            elif longest > 25:
                ws.row_dimensions[row].height = 30
            else:
                ws.row_dimensions[row].height = 20

            row += 1

    # v1.1: freeze first column AND header row so the vendor name stays
    # visible when scrolling through the 13-column wide row.
    # (Note: the per-row height logic above sizes rows by content length —
    # a previous unconditional `height = 40` loop here was overwriting that
    # sizing and has been removed.)
    ws.freeze_panes = "B5"

    # Print: landscape, fit to 1 page wide, unlimited pages tall.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "4:4"  # repeat header row on each printed page


# ---------------------------------------------------------------------------
# Sheet 3 — All Transactions
# ---------------------------------------------------------------------------

ALL_TRANSACTIONS_COLUMNS = [
    ("Source Statement", 40, "left"),
    ("Date",             12, "center"),
    ("Raw Description",  42, "left"),
    ("Canonical Vendor", 32, "left"),
    ("Amount ($)",       14, "right"),
    ("Excluded?",        10, "center"),
    ("Exclusion Reason", 28, "left"),
]


def write_all_transactions(
    wb: Workbook,
    agent_outputs: list[dict],
    filename_map: dict[str, str] | None = None,
):
    ws = wb.create_sheet("All Transactions")
    resolve = _make_resolver(filename_map)

    ws["A1"] = "ALL TRANSACTIONS — Combined Across Statements"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(ALL_TRANSACTIONS_COLUMNS))

    for col_idx, (header, width, _) in enumerate(ALL_TRANSACTIONS_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 26

    row = 4
    for out in agent_outputs:
        if out.get("status") not in ("success", "partial"):
            continue
        label = resolve(out["statement_label"])
        for t in out.get("transactions", []):
            row_data = [
                label,
                t.get("date", ""),
                t.get("raw_description", ""),
                t.get("canonical_name", ""),
                _round_currency(t.get("amount", 0.0)),
                "YES" if t.get("excluded") else "NO",
                t.get("exclusion_reason", ""),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.border = BORDER
                _, _, align = ALL_TRANSACTIONS_COLUMNS[col_idx - 1]
                cell.alignment = {"left": LEFT, "right": RIGHT, "center": CENTER}[align]
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            row += 1

    ws.freeze_panes = "A4"

    # Print: landscape, fit to 1 page wide.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "3:3"


# ---------------------------------------------------------------------------
# Sheet 4 — Per-Agent Summary
# ---------------------------------------------------------------------------

AGENT_SUMMARY_COLUMNS = [
    ("Statement",               40, "left"),
    ("Status",                  16, "center"),
    ("Transactions",            12, "center"),
    ("Vendors",                 10, "center"),
    ("Total Reconciled ($)",    18, "right"),
    ("Vendors over $600",       16, "center"),
    ("Review Needed",           14, "center"),
    ("Extraction Confidence",   18, "center"),
    ("Tool Calls",              10, "center"),
    ("Cost ($)",                10, "right"),
    ("Notes",                   40, "left"),
]


def write_per_agent_summary(
    wb: Workbook,
    agent_outputs: list[dict],
    filename_map: dict[str, str] | None = None,
):
    ws = wb.create_sheet("Per-Agent Summary")
    resolve = _make_resolver(filename_map)

    ws["A1"] = "PER-AGENT PROCESSING SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(AGENT_SUMMARY_COLUMNS))

    for col_idx, (header, width, _) in enumerate(AGENT_SUMMARY_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 28

    row = 4
    for out in agent_outputs:
        status = out.get("status", "unknown")
        is_success = status in ("success", "partial")

        if is_success:
            vendors = out.get("vendors", [])
            txn_count = sum(v.get("transaction_count", 0) for v in vendors)
            total = sum(v.get("total_amount", 0.0) for v in vendors)
            over_600 = sum(1 for v in vendors if v.get("total_amount", 0) >= 600)
            review = sum(1 for v in vendors if v.get("needs_review"))
        else:
            txn_count = 0
            vendors = []
            total = 0.0
            over_600 = 0
            review = 0

        row_data = [
            resolve(out["statement_label"]),
            status.upper().replace("_", " "),
            txn_count,
            len(vendors),
            _round_currency(total),
            over_600,
            review,
            out.get("extraction_confidence", 1.0) if is_success else 0.0,
            out.get("tool_calls", 0),
            _round_currency(out.get("cost_usd", 0.0)),
            out.get("error_message", "") or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            _, _, align = AGENT_SUMMARY_COLUMNS[col_idx - 1]
            cell.alignment = {"left": LEFT, "right": RIGHT, "center": CENTER}[align]

        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=8).number_format = '0%'
        ws.cell(row=row, column=10).number_format = '$0.0000'

        if status == "success":
            ws.cell(row=row, column=2).fill = ELIGIBLE_FILL
        elif status == "partial":
            ws.cell(row=row, column=2).fill = REVIEW_FILL
        elif status.startswith("failed"):
            ws.cell(row=row, column=2).fill = FAILED_FILL

        row += 1

    ws.freeze_panes = "A4"

    # Print: landscape, fit to 1 page wide.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "3:3"


# ---------------------------------------------------------------------------
# Sheet 2 — Validation Report
# ---------------------------------------------------------------------------

def write_validation_report(
    wb: Workbook,
    validation: DeterministicValidation,
    filename_map: dict[str, str] | None = None,
):
    ws = wb.create_sheet("Validation Report")
    resolve = _make_resolver(filename_map)

    ws["A1"] = "CROSS-STATEMENT VALIDATION REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    # Global column widths chosen to work for all 4 sections on this sheet,
    # since column widths apply sheet-wide. Tuned for:
    #   • Vendor / Name columns A-D wide enough for 25-char vendor names
    #     and statement filenames (e.g., "sample_credit_card_chase.pdf")
    #   • Numeric columns C, E generous for currency with commas
    #   • Last column F room for amounts, similarity, or variance ratios
    column_widths = [32, 32, 20, 32, 20, 24]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Red-bold font for high-variance cells (>3.0x ratio in Discrepancy Alerts)
    variance_critical_font = Font(name="Arial", bold=True, size=11, color="DC2626")

    row = 3

    def section_header(text: str, current_row: int) -> int:
        ws.cell(row=current_row, column=1, value=text).font = Font(
            name="Arial", bold=True, size=12, color="1F3A5F"
        )
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=6)
        return current_row + 1

    # ── Cross-Statement Vendor Matches ──
    row = section_header("Cross-Statement Vendor Matches", row)
    if validation.cross_matches:
        # 5-column table: Vendor / Statements / Combined Total / Crosses $600 /
        # Per-Statement Breakdown. The Per-Statement Breakdown column is the
        # full per-statement amount detail; no need for a separate footer.
        headers = ["Vendor", "Statements", "Combined Total ($)",
                   "Crosses $600 (Combined Only)", "Per-Statement Breakdown"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        row += 1
        for cm in validation.cross_matches:
            breakdown = "; ".join(
                f"{resolve(a['statement'])}: ${_round_currency(a['amount']):,.2f}"
                for a in cm.appearances
            )
            ws.cell(row=row, column=1, value=cm.canonical_name).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2, value=len(cm.appearances)).alignment = CENTER
            ws.cell(row=row, column=3, value=_round_currency(cm.combined_total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).alignment = RIGHT
            ws.cell(row=row, column=4,
                    value="YES — flag for filing" if cm.crosses_threshold_combined_only
                          else "no").alignment = CENTER
            ws.cell(row=row, column=5, value=breakdown).font = BODY_FONT
            ws.cell(row=row, column=5).alignment = LEFT
            if cm.crosses_threshold_combined_only:
                for c in range(1, 6):
                    ws.cell(row=row, column=c).fill = ELIGIBLE_FILL
            row += 1
    else:
        ws.cell(row=row, column=1,
                value="No vendors appear across multiple statements.").font = SUBTLE_FONT
        row += 1
    row += 2

    # ── Name Variant Flags ──
    row = section_header("Name Variant Flags", row)
    if validation.name_variants:
        headers = ["Statement A", "Name in A", "Statement B", "Name in B",
                   "Similarity", "Amounts"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        row += 1
        for nv in validation.name_variants:
            ws.cell(row=row, column=1, value=resolve(nv.statement_a)).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT  # wrap long filenames
            ws.cell(row=row, column=2, value=nv.name_a).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=3, value=resolve(nv.statement_b)).font = BODY_FONT
            ws.cell(row=row, column=3).alignment = LEFT
            ws.cell(row=row, column=4, value=nv.name_b).font = BODY_FONT
            ws.cell(row=row, column=4).alignment = LEFT
            ws.cell(row=row, column=5, value=nv.similarity).number_format = '0%'
            ws.cell(row=row, column=5).alignment = CENTER
            ws.cell(row=row, column=6,
                    value=f"${_round_currency(nv.amount_a):,.2f} / ${_round_currency(nv.amount_b):,.2f}").font = BODY_FONT
            ws.cell(row=row, column=6).alignment = LEFT
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = REVIEW_FILL
            row += 1
    else:
        ws.cell(row=row, column=1,
                value="No name variants detected across statements.").font = SUBTLE_FONT
        row += 1
    row += 2

    # ── Discrepancy Alerts ──
    row = section_header("Discrepancy Alerts — Possible Extraction Issues", row)
    if validation.amount_mismatches:
        headers = ["Vendor", "Statement A", "Amount A", "Statement B", "Amount B", "Variance"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        row += 1
        for am in validation.amount_mismatches:
            ws.cell(row=row, column=1, value=am.canonical_name).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2, value=resolve(am.statement_a)).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=3, value=_round_currency(am.amount_a)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).alignment = RIGHT
            ws.cell(row=row, column=4, value=resolve(am.statement_b)).font = BODY_FONT
            ws.cell(row=row, column=4).alignment = LEFT
            ws.cell(row=row, column=5, value=_round_currency(am.amount_b)).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).alignment = RIGHT
            # Variance: red bold when ratio > 3.0x to draw the eye to severe
            # extraction discrepancies; default styling otherwise.
            variance_cell = ws.cell(row=row, column=6, value=f"{am.ratio:.1f}x")
            variance_cell.alignment = CENTER
            if am.ratio > 3.0:
                variance_cell.font = variance_critical_font
            else:
                variance_cell.font = BODY_FONT
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = REVIEW_FILL
            row += 1
    else:
        ws.cell(row=row, column=1,
                value="No suspicious amount variances detected.").font = SUBTLE_FONT
        row += 1
    row += 2

    # ── Near-Threshold Vendors ──
    row = section_header("Near-Threshold Vendors ($500–$700 Review Zone)", row)
    if validation.near_threshold:
        headers = ["Vendor", "Source", "Total ($)", "Distance to $600"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        row += 1
        for nt in validation.near_threshold:
            ws.cell(row=row, column=1, value=nt.canonical_name).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2, value=resolve(nt.statement)).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=3, value=_round_currency(nt.total_amount)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3).alignment = RIGHT
            sign = "+" if nt.distance_to_threshold < 0 else "-"
            ws.cell(row=row, column=4,
                    value=f"{sign}${abs(_round_currency(nt.distance_to_threshold)):,.2f}").alignment = CENTER
            row += 1
    else:
        ws.cell(row=row, column=1,
                value="No vendors in the near-threshold review zone.").font = SUBTLE_FONT
        row += 1

    # Print: landscape, fit to 1 page wide.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ---------------------------------------------------------------------------
# Failed files log (defined but not invoked — preserved from v1.0)
# ---------------------------------------------------------------------------

def write_failed_files_log(
    wb: Workbook,
    agent_outputs: list[dict],
    filename_map: dict[str, str] | None = None,
):
    ws = wb.create_sheet("Failed Files Log")
    resolve = _make_resolver(filename_map)

    ws["A1"] = "FAILED / PARTIAL FILES LOG"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["Statement", "Status", "Reason", "Tool Calls Before Failure", "Cost Incurred ($)"]
    widths = [40, 18, 50, 24, 18]
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 26

    failed = [o for o in agent_outputs if o.get("status") not in ("success",)]

    if not failed:
        ws.cell(row=4, column=1,
                value="✓ All statements processed successfully — no failures.").font = BODY_FONT
        ws.cell(row=4, column=1).fill = ELIGIBLE_FILL
        ws.merge_cells("A4:E4")
        return

    row = 4
    for out in failed:
        status = out.get("status", "unknown")
        ws.cell(row=row, column=1, value=resolve(out["statement_label"])).font = BODY_FONT
        ws.cell(row=row, column=2,
                value=status.upper().replace("_", " ")).alignment = CENTER
        ws.cell(row=row, column=3,
                value=out.get("error_message") or "No vendor data extracted").font = BODY_FONT
        ws.cell(row=row, column=3).alignment = LEFT
        ws.cell(row=row, column=4, value=out.get("tool_calls", 0)).alignment = CENTER
        ws.cell(row=row, column=5, value=_round_currency(out.get("cost_usd", 0.0))).number_format = '$0.0000'

        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if status == "partial":
                cell.fill = REVIEW_FILL
            else:
                cell.fill = FAILED_FILL
        row += 1


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def generate_master_workbook(
    output_path: str,
    agent_outputs: list[dict],
    flags_by_statement: dict[str, dict[str, ReviewFlags]],
    eligibility_by_statement: dict[str, dict],
    validation: DeterministicValidation,
    filename_map: dict[str, str] | None = None,
) -> str:
    """
    Generate the consolidated 5-sheet master workbook (v1.1).

    Sheet order:
      0. Executive Summary    ★ new in v1.1
      1. Master Vendor Summary
      2. Validation Report
      3. All Transactions
      4. Per-Agent Summary

    Args:
        output_path:                path to write .xlsx
        agent_outputs:              list of agent result dicts
        flags_by_statement:         {statement_label: {vendor_name: ReviewFlags}}
        eligibility_by_statement:   {statement_label: {vendor_name: EligibilityResult}}
        validation:                 DeterministicValidation result
        filename_map:               optional {file_id: original_filename} mapping

    Returns:
        output_path
    """
    wb = Workbook()
    wb.remove(wb.active)

    # NEW in v1.1: Executive Summary first, becomes the active sheet on open
    write_executive_summary(
        wb, agent_outputs, flags_by_statement, eligibility_by_statement,
        validation, filename_map=filename_map,
    )
    write_master_vendor_summary(
        wb, agent_outputs, flags_by_statement, eligibility_by_statement,
        validation, filename_map=filename_map,
    )
    write_validation_report(wb, validation, filename_map=filename_map)
    write_all_transactions(wb, agent_outputs, filename_map=filename_map)
    write_per_agent_summary(wb, agent_outputs, filename_map=filename_map)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path