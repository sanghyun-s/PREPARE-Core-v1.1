"""
Server v1.0 — PREPARE Reconciliation Workspace
-----------------------------------------------
Single-page workflow with one main endpoint that dispatches to the chosen
engine, runs deterministic validation, generates the master Excel, and
returns one consolidated response built against the schemas.py contract.

v1.0 changes from v0.6:
    * AI Validation Narrative call removed. `run_validation_narrative` no
      longer imported or called.
    * `validation_model` form parameter dropped (clean cut). The dropdown is
      gone from the upload screen and the value is no longer carried.
    * `validation_models_available` field dropped from /api/health.
    * Per-agent `agent_narrative` field dropped from the response.
    * Response payload now conforms to schemas.py (ProcessResponse) with a
      formal contract — schema_version, summary, statements, validation,
      workbook, technical, errors. Frontend renders strictly against this
      shape.
    * Original filename preservation: uploads now retain `file.filename` and
      a per-run `filename_map` resolves uuid-based file IDs to readable
      filenames at the response boundary and inside the workbook.
    * `language` parameter retained as a no-op pass-through (carried for
      v1.1 workbook localization).

v1.3 changes:
    * New "pdf_skill" engine routed through pdf_skill_adapter.py via
      pipeline.run_pipeline_pdf_skill(). Existing engines (rule_based,
      ai_assisted, multi_agent) unchanged.
    * PDF Skill failures don't crash the run — they surface as per-statement
      "failed" entries with structured failure_reason.

Endpoints:
  POST /api/process              — Main workflow. Returns ProcessResponse.
  GET  /api/download/{file_id}   — Download a generated Excel file.
  GET  /api/health               — Health check.
"""

import os
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, NamedTuple

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv

# Path setup
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "backend"))

load_dotenv(ROOT / ".env")

from agent_app import run_engine, EXTRACTION_MODEL_DEFAULT
from review_flag_engine import build_flags_for_statement
from validation_engine import run_deterministic_validation
from master_excel_generator import generate_master_workbook
from vendor_classifier_1099 import classify_all_vendors

from schemas import (
    SCHEMA_VERSION,
    ProcessResponse,
    Summary,
    Statement,
    ReconciliationSnapshot,
    ExtractionCheck,
    Validation,
    CrossMatch,
    CrossMatchAppearance,
    NameVariant,
    DiscrepancyAlert,
    NearThresholdVendor,
    StatementExclusion,
    WorkbookInfo,
    Technical,
    RunError,
    HealthResponse,
)


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="PREPARE Reconciliation Workspace v1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOADS_DIR = ROOT / "uploads_tmp"
OUTPUTS_DIR = ROOT / "outputs"
UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

FRONTEND = ROOT / "frontend"

VERSION = "1.3.0"


@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the unified workspace UI."""
    index_path = FRONTEND / "index.html"
    if index_path.exists():
        return HTMLResponse(index_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Frontend not yet built</h1>")


@app.get("/api/health", response_model=HealthResponse)
async def health():
    return HealthResponse(
        version=VERSION,
        api_key_set=bool(os.getenv("ANTHROPIC_API_KEY")),
        extraction_model=EXTRACTION_MODEL_DEFAULT,
    )


# ---------------------------------------------------------------------------
# Upload helper — preserves original filename
# ---------------------------------------------------------------------------

class UploadInfo(NamedTuple):
    saved_path: Path
    original_filename: str
    file_id: str           # uuid stem (no extension), the canonical internal ID


def _save_upload(file: UploadFile, suffix: str) -> Optional[UploadInfo]:
    """
    Save uploaded file to uploads_tmp/ with a unique uuid-based name.
    Returns UploadInfo carrying the saved path, the user's original filename,
    and a stable file_id (uuid stem) for backend keying.
    """
    if not file:
        return None
    file_id = uuid.uuid4().hex
    saved_path = UPLOADS_DIR / f"{file_id}{suffix}"
    with open(saved_path, "wb") as f:
        f.write(file.file.read())
    return UploadInfo(
        saved_path=saved_path,
        original_filename=file.filename or saved_path.name,
        file_id=file_id,
    )


# ---------------------------------------------------------------------------
# v1.3: PDF Skill result → vendors list helper
# ---------------------------------------------------------------------------

def _vendors_from_pipeline_result(stmt_result: dict) -> list[dict]:
    """
    Convert the aggregated vendor summaries from run_pipeline_pdf_skill
    into the dict shape that downstream code (validation, response builder)
    expects.

    pipeline.run_pipeline_pdf_skill returns a 'vendor_summaries' list of
    VendorSummary objects in stmt_result. Each gets serialized to a dict
    matching the existing agent_outputs vendor schema.
    """
    summaries = stmt_result.get("vendor_summaries", [])
    if not summaries:
        return []

    return [
        {
            "canonical_name": s.canonical_name,
            "entity_type": s.entity_type,
            "total_amount": s.total_amount,
            "transaction_count": s.transaction_count,
            "match_confidence": s.match_confidence,
            "needs_review": s.needs_review,
            "review_reasons": list(s.review_reasons) if s.review_reasons else [],
            "raw_name_variants": list(s.raw_name_variants) if s.raw_name_variants else [],
        }
        for s in summaries
    ]


# ---------------------------------------------------------------------------
# Eligibility helper (unchanged from v0.6, just relocated)
# ---------------------------------------------------------------------------

def _build_eligibility_for_statement(vendors_dicts: list[dict]) -> dict:
    """Run the 1099 classifier on vendor dicts. Returns name -> EligibilityResult."""
    from dataclasses import dataclass, field

    @dataclass
    class _MinimalSummary:
        canonical_name: str
        entity_type: Optional[str]
        total_amount: float
        transaction_count: int = 1
        first_payment_date: Optional[str] = None
        last_payment_date: Optional[str] = None
        match_confidence: float = 1.0
        needs_review: bool = False
        review_reasons: list = field(default_factory=list)
        raw_name_variants: list = field(default_factory=list)

    mocks = []
    for v in vendors_dicts:
        mocks.append(_MinimalSummary(
            canonical_name=v["canonical_name"],
            entity_type=v.get("entity_type"),
            total_amount=v.get("total_amount", 0.0),
            transaction_count=v.get("transaction_count", 1),
            match_confidence=v.get("match_confidence", 1.0),
            needs_review=v.get("needs_review", False),
        ))
    return classify_all_vendors(mocks)


# ---------------------------------------------------------------------------
# Filename resolution
# ---------------------------------------------------------------------------

def _make_resolver(filename_map: dict[str, str]):
    """
    Build a label-resolution function. Given an internal label like
    "<file_id>.pdf" or just "<file_id>", returns the original filename.
    Passes through "Combined" and any unknown labels unchanged.
    """
    def resolve(label: str) -> str:
        if not label or label == "Combined":
            return label
        if label.endswith(".pdf"):
            stem = label[:-4]
        else:
            stem = label
        return filename_map.get(stem, label)
    return resolve


# ---------------------------------------------------------------------------
# Status mapping — collapse granular failure codes to schema's three-state status
# ---------------------------------------------------------------------------

def _map_status(raw_status: str) -> tuple[str, Optional[str]]:
    """
    Map agent_app's granular status to (schema_status, failure_reason).

    schema_status ∈ {"success", "partial", "failed"}
    failure_reason ∈ {"rate_limit", "extraction", "other", None}
    """
    if raw_status == "success":
        return "success", None
    if raw_status == "partial":
        return "partial", None
    if raw_status == "failed_rate_limit":
        return "failed", "rate_limit"
    if raw_status == "failed_extraction":
        return "failed", "extraction"
    return "failed", "other"


# ---------------------------------------------------------------------------
# Main endpoint — unified workflow
# ---------------------------------------------------------------------------

@app.post("/api/process", response_model=ProcessResponse)
async def process(
    pdf_files: list[UploadFile] = File(...),
    vendor_list: Optional[UploadFile] = File(None),
    engine: str = Form("multi_agent"),
    extraction_model: str = Form(EXTRACTION_MODEL_DEFAULT),
    language: str = Form("English"),
):
    """
    Unified workflow endpoint.

    engine:           "rule_based" | "ai_assisted" | "multi_agent" | "pdf_skill"
    extraction_model: Default Haiku. Advanced override available.
    language:         Reserved for v1.1 workbook localization. No-op in v1.0.
    """
    # ── Validate inputs ──
    if not pdf_files or len(pdf_files) == 0:
        raise HTTPException(status_code=422, detail="At least one PDF required.")
    if len(pdf_files) > 10:
        raise HTTPException(status_code=422, detail="Maximum 10 PDFs per run.")

    if engine not in ("rule_based", "ai_assisted", "multi_agent", "pdf_skill"):
        raise HTTPException(status_code=422, detail=f"Unknown engine: {engine}")

    if engine != "rule_based" and not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(
            status_code=400,
            detail="ANTHROPIC_API_KEY not set. Use rule_based engine or set the key in .env.",
        )

    t_start = time.time()

    # ── Save uploads, build filename_map ──
    upload_infos = [_save_upload(f, ".pdf") for f in pdf_files]
    pdf_paths = [str(u.saved_path) for u in upload_infos]
    filename_map: dict[str, str] = {u.file_id: u.original_filename for u in upload_infos}

    csv_info = _save_upload(vendor_list, ".csv") if vendor_list else None
    csv_path = str(csv_info.saved_path) if csv_info else None

    # ── Per-run output directory ──
    run_id = uuid.uuid4().hex[:12]
    run_dir = OUTPUTS_DIR / f"run_{run_id}"
    run_dir.mkdir(parents=True, exist_ok=True)

    run_errors: list[RunError] = []
    resolve = _make_resolver(filename_map)

    # ── Run the chosen engine ──
    # v1.3: PDF Skill engine bypasses run_engine() and uses pdf_skill_adapter
    # directly. It produces the same agent_outputs shape so downstream code
    # (validation, workbook, response building) stays identical.
    if engine == "pdf_skill":
        from pipeline import run_pipeline_pdf_skill

        agent_outputs = []
        total_cost_usd = 0.0
        any_success = False

        # Resolve model — only Sonnet and Opus are supported by pdf_skill_adapter
        pdf_skill_model = extraction_model if extraction_model in (
            "claude-sonnet-4-6", "claude-opus-4-7"
        ) else "claude-sonnet-4-6"

        for u in upload_infos:
            per_stmt_xlsx = run_dir / f"{u.file_id}_pdf_skill.xlsx"

            stmt_result = run_pipeline_pdf_skill(
                pdf_path=str(u.saved_path),
                output_path=str(per_stmt_xlsx),
                vendor_list_path=csv_path,
                source="bank",
                model=pdf_skill_model,
                verbose=True,
            )

            internal_label = f"{u.file_id}.pdf"

            if stmt_result.get("success"):
                any_success = True
                total_cost_usd += stmt_result.get("cost_usd", 0.0)
                agent_outputs.append({
                    "statement_label": internal_label,
                    "status": "success",
                    "error_message": None,
                    "vendors": _vendors_from_pipeline_result(stmt_result),
                    # v1.3 master-fix: pass real transactions through so master
                    # generator's All Transactions sheet and Executive Summary
                    # count have data to read. Falls back to empty list for
                    # safety if pipeline didn't expose them.
                    "transactions": stmt_result.get("transactions") or [],
                    "extraction_confidence": stmt_result.get("confidence", 0.0),
                    "tool_calls": 1,
                    # v1.3 master-fix: surface per-statement cost so master's
                    # Per-Agent Summary "Cost ($)" column renders the real number.
                    "cost_usd": stmt_result.get("cost_usd", 0.0),
                    "output_path": str(per_stmt_xlsx) if per_stmt_xlsx.exists() else None,
                    # v1.3 extras (surfaced in response for per-statement card)
                    "engine_used": "pdf_skill",
                    "pdf_skill_breakdown": stmt_result.get("pdf_skill_breakdown", {}),
                    "transactions_excluded": stmt_result.get("transactions_excluded", 0),
                    "pdf_skill_metadata": stmt_result.get("pdf_skill_metadata", {}),
                    # v1.4 Phase 4 — carry the computed reconciliation snapshot
                    # downstream. pipeline.run_pipeline_pdf_skill computes this
                    # via _compute_reconciliation() and puts it on its return
                    # dict; without this line it gets dropped here, leaving the
                    # per-statement card's waterfall and the Workspace 4E roll-up
                    # with no data (visible as "Unavailable" for every statement).
                    "reconciliation_snapshot": stmt_result.get("reconciliation_snapshot"),
                    # v1.4 Phase 4 — Source B: carry the extraction-completeness
                    # check downstream. Computed in pipeline._compute_source_b
                    # alongside _compute_reconciliation; same carry-or-drop
                    # pattern as the snapshot above. Without this line, the
                    # master 4E roll-up's "X of N show complete extraction"
                    # summary line wouldn't render and Statement.extraction_check
                    # would always be None.
                    "extraction_check": stmt_result.get("extraction_check"),
                })
            else:
                # PDF Skill failure — translate adapter failure_reason to
                # schema's raw status names so _map_status() handles it.
                failure_reason = stmt_result.get("failure_reason", "other")
                err_msg = stmt_result.get("error_details") or stmt_result.get(
                    "error", "PDF Skill extraction failed"
                )
                raw_status_map = {
                    "invalid_pdf":             "failed_extraction",
                    "agent_subprocess_failed": "failed_other",
                    "agent_timeout":           "failed_other",
                    "agent_returned_unknown":  "failed_extraction",
                    "schema_violation":        "failed_extraction",
                    "sdk_not_installed":       "failed_other",
                    "no_api_key":              "failed_other",
                }
                raw_status = raw_status_map.get(failure_reason, "failed_other")
                agent_outputs.append({
                    "statement_label": internal_label,
                    "status": raw_status,
                    "error_message": err_msg,
                    "vendors": [],
                    "transactions": [],
                    "extraction_confidence": 0.0,
                    "tool_calls": 0,
                    "cost_usd": 0.0,
                    "output_path": None,
                    "engine_used": "pdf_skill",
                    "pdf_skill_breakdown": {},
                    "transactions_excluded": 0,
                    "pdf_skill_metadata": {},
                })

        engine_result = {
            "success": True,  # Always continue downstream; per-statement failures are surfaced individually
            "agent_outputs": agent_outputs,
            "total_cost_usd": total_cost_usd,
        }

    else:
        # Existing engines (rule_based / ai_assisted / multi_agent)
        try:
            engine_result = await run_engine(
                engine=engine,
                pdf_paths=pdf_paths,
                output_dir=str(run_dir),
                vendor_list_path=csv_path,
                extraction_model=extraction_model,
                language=language,
            )
        except Exception as e:
            # Engine startup failure — top-level error, return minimal response
            raise HTTPException(status_code=500, detail=f"Engine failed: {e}")

        if not engine_result.get("success"):
            raise HTTPException(status_code=500, detail=engine_result.get("error", "Engine failed"))

        agent_outputs = engine_result.get("agent_outputs", [])

    # ── Build per-statement review flags + 1099 eligibility ──
    flags_by_statement: dict[str, dict] = {}
    eligibility_by_statement: dict[str, dict] = {}

    for out in agent_outputs:
        statement_label = out["statement_label"]   # internal uuid-based label

        if out.get("status") not in ("success", "partial"):
            flags_by_statement[statement_label] = {}
            eligibility_by_statement[statement_label] = {}
            continue

        from dataclasses import dataclass, field

        @dataclass
        class _MinimalSummary:
            canonical_name: str
            entity_type: Optional[str]
            total_amount: float
            transaction_count: int = 1
            first_payment_date: Optional[str] = None
            last_payment_date: Optional[str] = None
            match_confidence: float = 1.0
            needs_review: bool = False
            review_reasons: list = field(default_factory=list)
            raw_name_variants: list = field(default_factory=list)

        summaries = [
            _MinimalSummary(
                canonical_name=v["canonical_name"],
                entity_type=v.get("entity_type"),
                total_amount=v.get("total_amount", 0.0),
                transaction_count=v.get("transaction_count", 1),
                match_confidence=v.get("match_confidence", 1.0),
                needs_review=v.get("needs_review", False),
                review_reasons=v.get("review_reasons", []),
            )
            for v in out["vendors"]
        ]

        flags_by_statement[statement_label] = build_flags_for_statement(
            summaries,
            extraction_confidence=out.get("extraction_confidence", 1.0),
        )
        eligibility_by_statement[statement_label] = _build_eligibility_for_statement(
            out["vendors"]
        )

    # ── Run deterministic validation ──
    validation_raw = run_deterministic_validation(agent_outputs, flags_by_statement)

    # ── Generate master Excel workbook ──
    workbook_info: Optional[WorkbookInfo] = None
    workbook_filename = f"prepare_master_{datetime.now().strftime('%Y-%m-%d')}_{run_id}.xlsx"
    master_path = run_dir / workbook_filename

    try:
        generate_master_workbook(
            output_path=str(master_path),
            agent_outputs=agent_outputs,
            flags_by_statement=flags_by_statement,
            eligibility_by_statement=eligibility_by_statement,
            validation=validation_raw,
            filename_map=filename_map,
        )
        # Read actual sheet count from the generated workbook so the response
        # always reflects reality. v1.1 produces 5 sheets (Executive Summary
        # added at index 0); future sheet additions will be reflected
        # automatically without server.py edits.
        from openpyxl import load_workbook
        _wb_for_count = load_workbook(master_path, read_only=True)
        _actual_sheet_count = len(_wb_for_count.sheetnames)
        _wb_for_count.close()

        workbook_info = WorkbookInfo(
            file_id=f"run_{run_id}/{workbook_filename}",
            filename=workbook_filename,
            sheet_count=_actual_sheet_count,
            generated_at=datetime.now().isoformat(timespec="seconds"),
        )
    except Exception as e:
        run_errors.append(RunError(
            file_id=None,
            stage="workbook_generation",
            message=f"Master workbook generation failed: {e}",
            fatal=False,
        ))

    # ── Build per-agent file IDs for individual downloads ──
    for out in agent_outputs:
        if out.get("output_path"):
            p = Path(out["output_path"])
            if p.exists():
                out["_excel_file_id"] = f"run_{run_id}/{p.name}"
            else:
                out["_excel_file_id"] = None
        else:
            out["_excel_file_id"] = None

    # ── Build the response per ProcessResponse contract ──

    # Statements
    statements: list[Statement] = []
    successful_count = 0
    failed_count = 0
    total_transactions = 0
    total_vendors = 0
    total_amount = 0.0
    total_over_threshold = 0
    total_review_needed = 0

    # v1.4 Phase 4E (web tile) — cross-statement reconciliation roll-up.
    # Tally per-statement reconciliation status across SUCCESSFUL statements so
    # the Workspace can show "X of N reconcile" + balanced/needs_review/unavailable
    # breakdown alongside the other KPIs. A missing snapshot (rule_based /
    # multi_agent / no balance summary) folds into "unavailable" — same rule
    # as the master workbook 4E block.
    recon_balanced = 0
    recon_needs_review = 0
    recon_unavailable = 0

    for out in agent_outputs:
        raw_status = out.get("status", "failed_other")
        schema_status, failure_reason = _map_status(raw_status)

        is_ok = schema_status in ("success", "partial")
        if is_ok:
            successful_count += 1
        else:
            failed_count += 1

        vendors = out.get("vendors", []) if is_ok else []
        all_txns = out.get("transactions", []) if is_ok else []

        # v1.3: for pdf_skill engine, transaction_count comes from vendor summaries
        # since `transactions` list isn't surfaced to keep response payload small.
        if out.get("engine_used") == "pdf_skill" and is_ok:
            txn_count = sum(v.get("transaction_count", 0) for v in vendors)
        else:
            txn_count = sum(1 for t in all_txns if t.get("include_for_1099", True))

        amt = round(sum(v.get("total_amount", 0) for v in vendors), 2)
        over_thresh = sum(1 for v in vendors if v.get("total_amount", 0) >= 600)
        review_count = sum(
            1 for f in flags_by_statement.get(out["statement_label"], {}).values()
            if f.needs_review
        )

        total_transactions += txn_count
        total_vendors += len(vendors)
        total_amount += amt
        total_over_threshold += over_thresh
        total_review_needed += review_count

        # file_id from internal label (strip .pdf)
        internal_label = out["statement_label"]
        file_id_stem = internal_label[:-4] if internal_label.endswith(".pdf") else internal_label

        # v1.4 (Phase 4): build the reconciliation snapshot model from the dict
        # the pipeline computed. None when the engine produced no snapshot
        # (rule_based / multi_agent) or the statement failed.
        recon_dict = out.get("reconciliation_snapshot")
        recon_model = (
            ReconciliationSnapshot(**recon_dict)
            if isinstance(recon_dict, dict) and recon_dict
            else None
        )

        # v1.4 (Phase 4 — Source B): build the extraction-completeness model
        # from the dict the pipeline computed. None when the engine produced
        # no check (rule_based / multi_agent) or the statement failed. Same
        # carry-or-drop pattern as recon_model above.
        ec_dict = out.get("extraction_check")
        ec_model = (
            ExtractionCheck(**ec_dict)
            if isinstance(ec_dict, dict) and ec_dict
            else None
        )

        # v1.4 Phase 4E (web tile) — tally for the roll-up, successful only.
        if is_ok:
            status = recon_dict.get("status") if isinstance(recon_dict, dict) else None
            if status == "balanced":
                recon_balanced += 1
            elif status == "needs_review":
                recon_needs_review += 1
            else:
                recon_unavailable += 1

        statements.append(Statement(
            file_id=file_id_stem,
            original_filename=resolve(internal_label),
            status=schema_status,
            failure_reason=failure_reason,
            error_message=out.get("error_message"),
            transaction_count=txn_count,
            vendor_count=len(vendors),
            total_amount=amt,
            vendors_over_threshold=over_thresh,
            review_needed=review_count,
            extraction_confidence=out.get("extraction_confidence", 0.0),
            excel_file_id=out.get("_excel_file_id"),
            # v1.3: pass through PDF Skill fields if present (None for other engines)
            engine_used=out.get("engine_used"),
            bookkeeping_breakdown=out.get("pdf_skill_breakdown") or None,
            excluded_count=out.get("transactions_excluded", 0),
            # v1.4 (Phase 4): reconciliation snapshot (None for non-PDF-Skill engines)
            reconciliation_snapshot=recon_model,
            # v1.4 (Phase 4 — Source B): extraction-completeness check
            # (None for non-PDF-Skill engines or when no snapshot was available
            # to compare against).
            extraction_check=ec_model,
        ))

    # Summary — unique_vendors is the deduplicated count of canonical names
    # across all successful statements (matches Executive Summary workbook sheet).
    # Without dedup, vendors appearing in multiple statements get double-counted.
    _unique_vendor_names: set[str] = set()
    for _o in agent_outputs:
        if _o.get("status") in ("success", "partial"):
            for _v in _o.get("vendors", []):
                _unique_vendor_names.add(_v["canonical_name"])

    summary = Summary(
        total_transactions=total_transactions,
        unique_vendors=len(_unique_vendor_names),
        total_reconciled=round(total_amount, 2),
        vendors_over_threshold=total_over_threshold,
        review_needed_count=total_review_needed,
        successful_count=successful_count,
        failed_count=failed_count,
        # v1.4 Phase 4E (web tile) — cross-statement reconciliation roll-up
        reconciliation_balanced=recon_balanced,
        reconciliation_needs_review=recon_needs_review,
        reconciliation_unavailable=recon_unavailable,
    )

    # Validation — resolve all statement labels to original filenames
    validation = Validation(
        cross_statement_matches=[
            CrossMatch(
                vendor=cm.canonical_name,
                statement_count=len(cm.appearances),
                combined_total=round(cm.combined_total, 2),
                crosses_threshold_combined_only=cm.crosses_threshold_combined_only,
                appearances=[
                    CrossMatchAppearance(
                        statement=resolve(a["statement"]),
                        amount=round(a["amount"], 2),
                        count=a["count"],
                    )
                    for a in cm.appearances
                ],
            )
            for cm in validation_raw.cross_matches
        ],
        name_variants=[
            NameVariant(
                statement_a=resolve(nv.statement_a),
                name_a=nv.name_a,
                statement_b=resolve(nv.statement_b),
                name_b=nv.name_b,
                similarity=nv.similarity,
                amount_a=round(nv.amount_a, 2),
                amount_b=round(nv.amount_b, 2),
            )
            for nv in validation_raw.name_variants
        ],
        discrepancy_alerts=[
            DiscrepancyAlert(
                vendor=am.canonical_name,
                statement_a=resolve(am.statement_a),
                amount_a=round(am.amount_a, 2),
                statement_b=resolve(am.statement_b),
                amount_b=round(am.amount_b, 2),
                ratio=am.ratio,
                abs_diff=round(am.abs_diff, 2),
            )
            for am in validation_raw.amount_mismatches
        ],
        near_threshold_vendors=[
            NearThresholdVendor(
                vendor=nt.canonical_name,
                source=resolve(nt.statement),    # "Combined" passes through
                total_amount=round(nt.total_amount, 2),
                distance_to_threshold=round(nt.distance_to_threshold, 2),
            )
            for nt in validation_raw.near_threshold
        ],
        statements_processed=[resolve(s) for s in validation_raw.statements_processed],
        statements_excluded=[
            StatementExclusion(
                statement=resolve(ex["statement"]),
                status=ex["status"],
                reason=ex["reason"],
            )
            for ex in validation_raw.statements_excluded
        ],
    )

    # Technical
    total_tool_calls = sum(o.get("tool_calls", 0) for o in agent_outputs)
    total_cost = engine_result.get("total_cost_usd", 0.0)

    technical = Technical(
        engine=engine,
        extraction_model=extraction_model if engine != "rule_based" else None,
        language=language,
        processing_time_seconds=round(time.time() - t_start, 2),
        total_cost_usd=round(total_cost, 4),
        total_tool_calls=total_tool_calls,
    )

    return ProcessResponse(
        schema_version=SCHEMA_VERSION,
        success=True,
        run_id=run_id,
        summary=summary,
        statements=statements,
        validation=validation,
        workbook=workbook_info,
        technical=technical,
        errors=run_errors,
    )


# ---------------------------------------------------------------------------
# Download endpoint — supports nested run_*/file.xlsx paths
# ---------------------------------------------------------------------------

@app.get("/api/download/{file_id:path}")
async def download(file_id: str):
    """Download an Excel file by its file_id (which may include run subdir)."""
    safe = Path(file_id).as_posix()
    if ".." in safe or safe.startswith("/"):
        raise HTTPException(status_code=400, detail="Invalid file_id")

    full_path = OUTPUTS_DIR / safe
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {file_id}")

    return FileResponse(
        full_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=full_path.name,
    )


# ---------------------------------------------------------------------------
# Static frontend
# ---------------------------------------------------------------------------

if FRONTEND.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND)), name="static")